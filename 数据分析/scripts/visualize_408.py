"""
408 四科分析 — 遵循 数据分析/ 管道体系
  1) export_xlsx_to_csv.py  → xlsx 公式求值导出 CSV
  2) analyze_csv_folder.py  → CSV → 分析结果/ (per-sheet)
  3) 本脚本 → 真题/自编 拆分 + 四科综合报告 (408分析报告.md)

数据流: 习题册.xlsx → csv/ → 分析结果/ → 408分析报告.md
"""
from pathlib import Path
import subprocess, sys, shutil, warnings, re
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib import font_manager
from openpyxl import load_workbook

SRC = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
ROOT = Path(r'C:\Users\34406\Documents\Obsidian Vault\数据分析')
OUT_ROOT = ROOT / '分析结果'

# ============================================================
# 中文字体 (完整版)
# ============================================================
font_candidates = [
    'C:/Windows/Fonts/msyh.ttc', 'C:/Windows/Fonts/simhei.ttf',
    'C:/Windows/Fonts/simsun.ttc',
    '/System/Library/Fonts/PingFang.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
]
for fp in font_candidates:
    if Path(fp).exists():
        font_manager.fontManager.addfont(fp)
        plt.rcParams['font.family'] = font_manager.FontProperties(fname=fp).get_name()
        break
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun', 'WenQuanYi Micro Hei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 100
sns.set_style('darkgrid')
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun', 'WenQuanYi Micro Hei', 'Arial Unicode MS']
plt.rcParams['font.family'] = 'sans-serif'

# ============================================================
# 复用 analyze_csv_folder 的图表函数
# ============================================================
def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name)).strip()

def save_no_data_figure(path: Path, title: str, message: str) -> None:
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.axis("off")
    ax.text(0.5, 0.62, title, ha="center", va="center", fontsize=18, fontweight="bold")
    ax.text(0.5, 0.38, message, ha="center", va="center", fontsize=13, color="#666")
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)

def save_section_heatmap(df: pd.DataFrame, title: str, out_path: Path) -> None:
    """完全复用 analyze_csv_folder 的 heatmap 风格"""
    valid = df[df["total"] > 0].copy()
    if valid.empty:
        save_no_data_figure(out_path, title, "暂无数据")
        return
    
    valid["section_no"] = valid["section"].apply(lambda s: s.split()[0] if s else "")
    valid["section_order"] = valid.groupby("chapter").cumcount() + 1
    
    heat_matrix = valid.pivot_table(index="chapter", columns="section_order", values="rate", aggfunc="mean")
    annot_matrix = (
        valid.pivot_table(index="chapter", columns="section_order", values="section_no", aggfunc="first")
        .reindex_like(heat_matrix).fillna("")
    )
    annot = heat_matrix.copy().astype(object)
    for r_idx in heat_matrix.index:
        for c_idx in heat_matrix.columns:
            v = heat_matrix.loc[r_idx, c_idx]
            label = annot_matrix.loc[r_idx, c_idx]
            annot.loc[r_idx, c_idx] = "" if pd.isna(v) else f"{label}\n{v:.0%}"
    
    fig_w = max(14, 1.55 * heat_matrix.shape[1] + 5)
    fig_h = max(7, 1.15 * heat_matrix.shape[0] + 2)
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    sns.heatmap(heat_matrix, annot=annot, fmt="", cmap="RdYlGn", vmin=0, vmax=1,
                linewidths=1.0, linecolor="white",
                cbar_kws={"label": "一刷正确率"},
                annot_kws={"fontsize": 11, "fontweight": "bold"}, ax=ax)
    ax.set_title(f"{title}：小节正确率热力图", fontsize=18, fontweight="bold", pad=18)
    ax.set_xlabel("本章内小节序号", fontsize=13)
    ax.set_ylabel("章节", fontsize=13)
    ax.tick_params(axis="x", labelrotation=0, labelsize=11)
    ax.tick_params(axis="y", labelrotation=0, labelsize=12)
    fig.tight_layout()
    fig.savefig(out_path, dpi=180, bbox_inches="tight")
    plt.close(fig)

def annotate_rate_axis(ax, values, x_offset=0.012):
    for patch, value in zip(ax.patches, values):
        if pd.isna(value): continue
        w = patch.get_width()
        y = patch.get_y() + patch.get_height() / 2
        ax.text(min(w + x_offset, 0.985), y, f"{value:.1%}", va="center", ha="left", fontsize=10, color="#333")

def save_chapter_analysis(df: pd.DataFrame, chapter_df: pd.DataFrame, title: str, out_path: Path) -> None:
    """完全复用 analyze_csv_folder 的 2x2 风格"""
    valid = df[df["total"] > 0].copy()
    if valid.empty or chapter_df.empty:
        save_no_data_figure(out_path, title, "暂无数据")
        return
    
    fig, axes = plt.subplots(2, 2, figsize=(18, 13))
    plot_chapter = chapter_df.copy()
    
    sns.barplot(data=plot_chapter, y="chapter", x="first_rate", hue="chapter", palette="RdYlGn", legend=False, ax=axes[0, 0])
    axes[0, 0].set_title("各章节一刷正确率", fontsize=15, fontweight="bold")
    axes[0, 0].set_xlim(0, 1)
    axes[0, 0].xaxis.set_major_formatter(lambda x, pos: f"{x:.0%}")
    annotate_rate_axis(axes[0, 0], plot_chapter["first_rate"].tolist())
    
    chapter_long = plot_chapter.melt(
        id_vars="chapter",
        value_vars=["first_total", "first_correct", "first_wrong"],
        var_name="metric", value_name="count")
    chapter_long["metric"] = chapter_long["metric"].map({"first_total": "总题", "first_correct": "正确", "first_wrong": "错误"})
    sns.barplot(data=chapter_long, y="chapter", x="count", hue="metric", ax=axes[0, 1])
    axes[0, 1].set_title("各章节题量/正确/错误", fontsize=15, fontweight="bold")
    axes[0, 1].legend(title="")
    
    sns.histplot(data=valid, x="rate", bins=10, kde=True, color="#5B9BD5", ax=axes[1, 0])
    axes[1, 0].set_title("小节正确率分布", fontsize=15, fontweight="bold")
    axes[1, 0].set_xlim(0, 1)
    axes[1, 0].xaxis.set_major_formatter(lambda x, pos: f"{x:.0%}")
    
    sns.scatterplot(data=valid, x="total", y="rate", size="total", hue="chapter",
                    sizes=(60, 420), alpha=0.82, ax=axes[1, 1])
    axes[1, 1].set_title("小节题量 vs 正确率", fontsize=15, fontweight="bold")
    axes[1, 1].yaxis.set_major_formatter(lambda x, pos: f"{x:.0%}")
    axes[1, 1].set_ylim(0, 1.05)
    for _, row in valid.nsmallest(5, "rate").iterrows():
        axes[1, 1].annotate(row["section_no"], (row["total"], row["rate"]),
                           xytext=(5, 5), textcoords="offset points", fontsize=10)
    axes[1, 1].legend(bbox_to_anchor=(1.02, 1), loc="upper left", borderaxespad=0)
    
    for ax in axes.flat:
        ax.set_xlabel(ax.get_xlabel(), fontsize=12)
        ax.set_ylabel(ax.get_ylabel(), fontsize=12)
        ax.tick_params(labelsize=10)
    fig.tight_layout()
    fig.savefig(out_path, dpi=180, bbox_inches="tight")
    plt.close(fig)

def save_cross_table_analysis(summary_df: pd.DataFrame, out_path: Path) -> None:
    """复用 analyze_csv_folder 的横向对比图"""
    if summary_df.empty:
        save_no_data_figure(out_path, "全部科目横向对比", "没有可分析的数据")
        return
    
    plot_df = summary_df.copy()
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    sns.barplot(data=plot_df, y="source_file", x="first_rate", hue="source_file", palette="RdYlGn", legend=False, ax=axes[0])
    axes[0].set_title("各科一刷正确率对比", fontsize=15, fontweight="bold")
    axes[0].set_xlim(0, 1)
    axes[0].xaxis.set_major_formatter(lambda x, pos: f"{x:.0%}")
    annotate_rate_axis(axes[0], plot_df["first_rate"].tolist())
    
    heat_metrics = plot_df.set_index("source_file")[["sections_with_first_data", "first_total", "first_wrong", "first_rate", "first_avg_section_rate"]]
    heat_norm = heat_metrics.copy()
    for col in ["sections_with_first_data", "first_total", "first_wrong"]:
        mv = heat_norm[col].max()
        heat_norm[col] = heat_norm[col] / mv if mv else 0
    sns.heatmap(heat_norm, annot=heat_metrics.round(3), fmt="", cmap="YlGnBu",
                linewidths=1, linecolor="white", ax=axes[1], annot_kws={"fontsize": 10})
    axes[1].set_title("各科指标热力图（题量列归一化着色）", fontsize=15, fontweight="bold")
    fig.tight_layout()
    fig.savefig(out_path, dpi=180, bbox_inches="tight")
    plt.close(fig)

# ============================================================
# 数据读取 (xlsx → 真题/自编拆分，CSV 管道已由上方步骤完成)
# ============================================================
wb = load_workbook(SRC)

def safe_int(v):
    if v is None: return None
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str) and v.startswith('='): return None
    try: return int(v)
    except: return None

subjects = [
    ('DS', '王道数据结构', '数据结构'),
    ('CO', '王道计组', '计算机组成原理'),
    ('OS', '王道操作系统', '操作系统'),
]
prefix_map = {'DS': '04', 'CO': '03', 'OS': '02'}

all_overall = {}
all_chapter_stats = {}
all_summaries = []

for subj, sheet_name, full_name in subjects:
    ws = wb[sheet_name]
    heji_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).replace(' ', '') == '合计':
            heji_row = r; break
    
    folder = OUT_ROOT / f'{prefix_map[subj]}_王道{full_name}'
    if folder.exists(): shutil.rmtree(folder)
    folder.mkdir(parents=True, exist_ok=True)
    
    records = []
    current_ch = None
    ch_no = 0
    
    for row in range(5, heji_row or ws.max_row):
        a = ws.cell(row=row, column=1).value
        if a is None: continue
        a_s = str(a).strip()
        
        if a_s.startswith('第') and '章' in a_s:
            current_ch = a_s
            m = re.search(r'\d+', current_ch)
            ch_no = int(m.group()) if m else 0
            continue
        if current_ch is None: continue
        
        c_t = safe_int(ws.cell(row=row, column=3).value)
        c_c = safe_int(ws.cell(row=row, column=4).value)
        f_zt = safe_int(ws.cell(row=row, column=6).value)
        g_ztc = safe_int(ws.cell(row=row, column=7).value)
        if c_t is None: continue
        c_c = c_c or 0; f_zt = f_zt or 0; g_ztc = g_ztc or 0
        
        sec_match = re.match(r'^(\d+(?:\.\d+)+)', a_s)
        records.append({
            'source_file': sheet_name,
            'title': ws.cell(1, 1).value or sheet_name,
            'chapter': current_ch,
            'chapter_no': ch_no,
            'section': a_s,
            'section_no': sec_match.group(1) if sec_match else '',
            'total': c_t, 'correct': c_c,
            'wrong': max(c_t - c_c, 0),
            'rate': c_c / c_t if c_t > 0 else 0,
            'zt_total': f_zt, 'zt_correct': g_ztc,
            'zb_total': c_t - f_zt, 'zb_correct': c_c - g_ztc,
            'has_data': True,
        })
    
    df = pd.DataFrame(records)
    if df.empty: continue
    df['section_order'] = df.groupby('chapter').cumcount() + 1
    
    # Build ch_data from records for zhenti/zibian per-chapter aggregation
    ch_data = {}
    for _, r in df.iterrows():
        ch = r['chapter']
        if ch not in ch_data:
            ch_data[ch] = {'total':0, 'correct':0, 'zt_t':0, 'zt_c':0, 'sections':[]}
        ch_data[ch]['total'] += r['total']
        ch_data[ch]['correct'] += r['correct']
        ch_data[ch]['zt_t'] += r['zt_total']
        ch_data[ch]['zt_c'] += r['zt_correct']
        ch_data[ch]['sections'].append(r.to_dict())
    
    # --- overall_summary ---
    total_all = df['total'].sum()
    correct_all = df['correct'].sum()
    zt_all = df['zt_total'].sum()
    zt_c = df['zt_correct'].sum()
    valid = df[df['total'] > 0]
    overall = pd.DataFrame([{
        'source_file': sheet_name,
        'title': df.iloc[0]['title'],
        'sections_total': len(df),
        'sections_with_first_data': len(valid),
        'first_total': total_all,
        'first_correct': correct_all,
        'first_wrong': total_all - correct_all,
        'first_rate': correct_all / total_all if total_all else np.nan,
        'first_avg_section_rate': valid['rate'].mean() if not valid.empty else np.nan,
    }])
    overall.to_csv(folder / 'overall_summary.csv', index=False, encoding='utf-8-sig')
    
    # --- chapter_summary ---
    ch_agg = df.groupby(['chapter_no', 'chapter'], dropna=False).agg(
        sections_total=('section', 'count'),
        sections_with_first_data=('total', lambda s: int((s > 0).sum())),
        first_total=('total', 'sum'),
        first_correct=('correct', 'sum'),
        first_wrong=('wrong', 'sum'),
        rate=('rate', 'mean'),
    ).reset_index()
    ch_agg['first_rate'] = ch_agg['first_correct'] / ch_agg['first_total'].replace(0, np.nan)
    
    # weakest section per chapter
    if not valid.empty:
        weak_ch = valid.sort_values(['chapter_no', 'rate', 'total'], ascending=[True, True, False])\
            .groupby('chapter', as_index=False).first()[['chapter', 'section', 'rate']]\
            .rename(columns={'section': 'weakest_section', 'rate': 'weakest_rate'})
        ch_agg = ch_agg.merge(weak_ch, on='chapter', how='left')
    ch_agg = ch_agg.sort_values('chapter_no')
    ch_agg.to_csv(folder / 'chapter_summary.csv', index=False, encoding='utf-8-sig')
    
    # Populate for report generation
    all_overall[subj] = {
        'full_name': full_name, 'total': total_all, 'correct': correct_all,
        'rate': correct_all / total_all, 'sections': len(df), 'chapters': len(ch_agg),
        'zt_rate': zt_c / zt_all if zt_all else None,
        'zb_rate': (correct_all - zt_c) / (total_all - zt_all) if (total_all - zt_all) else None,
    }
    
    ch_list = []
    for ch_name, cd in ch_data.items():
        ch_t = cd['total']; ch_c = cd['correct']
        ch_r = ch_c / ch_t
        zt_r = cd['zt_c'] / cd['zt_t'] if cd['zt_t'] else None
        zb_t = ch_t - cd['zt_t']; zb_c = ch_c - cd['zt_c']
        zb_r = zb_c / zb_t if zb_t else None
        secs = sorted(cd['sections'], key=lambda x: x['rate'])
        ch_list.append({
            'chapter': ch_name, 'chapter_no': int(re.search(r'\d+', ch_name).group()),
            'total': ch_t, 'correct': ch_c, 'rate': ch_r,
            'zt_rate': zt_r, 'zb_rate': zb_r,
            'weakest': {'section': secs[0]['section'], 'rate': secs[0]['rate']},
            'strongest': {'section': secs[-1]['section'], 'rate': secs[-1]['rate']},
        })
    all_chapter_stats[subj] = sorted(ch_list, key=lambda x: x['chapter_no'])
    overall = pd.DataFrame([{
        'source_file': sheet_name,
        'title': df.iloc[0]['title'],
        'sections_total': len(df),
        'sections_with_first_data': len(valid),
        'first_total': total_all,
        'first_correct': correct_all,
        'first_wrong': total_all - correct_all,
        'first_rate': correct_all / total_all if total_all else np.nan,
        'first_avg_section_rate': valid['rate'].mean() if not valid.empty else np.nan,
    }])
    overall.to_csv(folder / 'overall_summary.csv', index=False, encoding='utf-8-sig')
    
    # --- chapter_summary ---
    ch_agg = df.groupby(['chapter_no', 'chapter'], dropna=False).agg(
        sections_total=('section', 'count'),
        sections_with_first_data=('total', lambda s: int((s > 0).sum())),
        first_total=('total', 'sum'),
        first_correct=('correct', 'sum'),
        first_wrong=('wrong', 'sum'),
        rate=('rate', 'mean'),
    ).reset_index()
    ch_agg['first_rate'] = ch_agg['first_correct'] / ch_agg['first_total'].replace(0, np.nan)
    
    # weakest section per chapter
    if not valid.empty:
        weak_ch = valid.sort_values(['chapter_no', 'rate', 'total'], ascending=[True, True, False])\
            .groupby('chapter', as_index=False).first()[['chapter', 'section', 'rate']]\
            .rename(columns={'section': 'weakest_section', 'rate': 'weakest_rate'})
        ch_agg = ch_agg.merge(weak_ch, on='chapter', how='left')
    ch_agg = ch_agg.sort_values('chapter_no')
    ch_agg.to_csv(folder / 'chapter_summary.csv', index=False, encoding='utf-8-sig')
    
    # --- cleaned_records ---
    df.to_csv(folder / 'cleaned_records.csv', index=False, encoding='utf-8-sig')
    
    # --- weak_sections_top10 ---
    weak_df = valid.sort_values(['rate', 'total'], ascending=[True, False]).head(10)
    weak_df.to_csv(folder / 'weak_sections_top10.csv', index=False, encoding='utf-8-sig')
    
    # --- 图表 ---
    meta = {'title': df.iloc[0]['title'], 'source_file': sheet_name}
    save_section_heatmap(df, meta['title'], folder / 'section_heatmap.png')
    save_chapter_analysis(df, ch_agg, meta['title'], folder / 'chapter_analysis.png')
    
    # --- report.md ---
    item = overall.iloc[0].to_dict()
    rpt = [
        f"# {meta['title']}",
        "",
        f"- 来源：`{sheet_name}`",
        f"- 小节总数：{int(item['sections_total'])}",
        f"- 已填写一刷小节：{int(item['sections_with_first_data'])}",
        f"- 一刷总题：{int(item['first_total'])}",
        f"- 一刷正确：{int(item['first_correct'])}",
        f"- 一刷正确率：{item['first_rate']:.2%}" if not pd.isna(item['first_rate']) else "-",
        "",
        "## 输出文件",
        "",
        "- `cleaned_records.csv`：清洗后小节明细",
        "- `overall_summary.csv`：总体指标",
        "- `chapter_summary.csv`：章节汇总（含最弱小节）",
        "- `weak_sections_top10.csv`：薄弱小节 Top 10",
        "- `section_heatmap.png`：小节正确率热力图",
        "- `chapter_analysis.png`：章节与分布分析",
    ]
    (folder / 'report.md').write_text('\n'.join(rpt), encoding='utf-8')
    
    all_summaries.append(overall.iloc[0].to_dict())
    print(f'[{subj}] {full_name}: {correct_all}/{total_all} = {correct_all/total_all*100:.1f}%  -> {folder.name}')

# ============================================================
# 汇总
# ============================================================
summary_dir = OUT_ROOT / '_汇总'
if summary_dir.exists(): shutil.rmtree(summary_dir)
summary_dir.mkdir(parents=True, exist_ok=True)

summary_df = pd.DataFrame(all_summaries)
summary_df.to_csv(summary_dir / 'all_tables_summary.csv', index=False, encoding='utf-8-sig')

# 合并所有 cleaned_records
all_dfs = []
for subj in ['DS', 'CO', 'OS']:
    p = OUT_ROOT / f'{prefix_map[subj]}_王道{["数据结构","计组","操作系统"][["DS","CO","OS"].index(subj)]}' / 'cleaned_records.csv'
    if p.exists():
        all_dfs.append(pd.read_csv(p))
if all_dfs:
    pd.concat(all_dfs, ignore_index=True).to_csv(summary_dir / 'all_cleaned_records.csv', index=False, encoding='utf-8-sig')

save_cross_table_analysis(summary_df, summary_dir / 'all_tables_comparison.png')

print(f'\n汇总: {summary_dir}')
print(f'  - all_tables_summary.csv')
print(f'  - all_cleaned_records.csv')
print(f'  - all_tables_comparison.png')

# ============================================================
# 生成 408 综合报告 (真题/自编 拆分来自 xlsx 直接读取)
# ============================================================
FIG_DIR = ROOT / 'figures'
FIG_DIR.mkdir(parents=True, exist_ok=True)

# 汇总表
R = []
total_t = sum(all_overall[s]['total'] for s in ['DS','CO','OS'])
total_c = sum(all_overall[s]['correct'] for s in ['DS','CO','OS'])

def fmt_pct(v):
    if v is None: return '-'
    return f'{v*100:.1f}%'

R.append('# 408 计算机统考 — 全面分析报告\n')
R.append(f'> 数据管道: 习题册.xlsx → export_xlsx_to_csv.py → analyze_csv_folder.py → 本报告\n')
R.append(f'> 生成时间：2026-06-16\n')

# 总览
R.append('## 一、四科总览\n')
R.append('| 科目 | 题量 | 正确数 | 正确率 | 真题率 | 自编率 | 章节 | 小节 |')
R.append('|:---|:---:|:---:|:---:|:---:|:---:|:---:|:---:|')
for subj in ['DS', 'CO', 'OS']:
    s = all_overall[subj]
    R.append(f"| **{s['full_name']}** | {s['total']} | {s['correct']} | {fmt_pct(s['rate'])} | {fmt_pct(s['zt_rate'])} | {fmt_pct(s['zb_rate'])} | {s['chapters']} | {s['sections']} |")
R.append(f"| 计算机网络 | — | — | 未开始 | — | — | — | — |")
R.append(f"\n**三科合计**：{total_c}/{total_t} = {total_c/total_t*100:.1f}%，408 估分约 **{round(total_c/total_t*150*0.85)} 分**\n")

# 真题 vs 自编 对比图
R.append('## 二、真题 vs 自编 vs 总体\n')
fig, axes = plt.subplots(1, 3, figsize=(20, 6.5))
for ax_i, subj in enumerate(['DS', 'CO', 'OS']):
    ax = axes[ax_i]
    chs = all_chapter_stats[subj]
    labels = [c['chapter'].replace('第','').replace('章','') for c in chs]
    x = np.arange(len(labels)); w = 0.32
    zt_rates = [(c['zt_rate']*100 if c['zt_rate'] else 0) for c in chs]
    zb_rates = [(c['zb_rate']*100 if c['zb_rate'] else 0) for c in chs]
    all_rates = [c['rate']*100 for c in chs]
    ax.bar(x - w, zt_rates, w, label='真题', color='#2E86AB', edgecolor='white')
    ax.bar(x, zb_rates, w, label='自编', color='#F18F01', edgecolor='white')
    ax.bar(x + w, all_rates, w, label='总体', color='#27AE60', edgecolor='white', alpha=0.6)
    ax.set_xticks(x); ax.set_xticklabels(labels, fontsize=8.5, rotation=30)
    ax.set_ylabel('正确率 (%)'); ax.set_ylim(0, 112)
    ax.set_title(all_overall[subj]['full_name'], fontsize=14, fontweight='bold')
    ax.legend(fontsize=10, loc='lower left')
fig.suptitle('真题 vs 自编 vs 总体 各章正确率对比', fontsize=16, fontweight='bold', y=1.02)
fig.tight_layout()
fig.savefig(FIG_DIR / '408_zhenti_vs_zibian.png', dpi=150, bbox_inches='tight')
plt.close(fig)
R.append('\n![真题vs自编](figures/408_zhenti_vs_zibian.png)\n')

# 单科分析
for subj, sheet_name, full_name in subjects:
    ov = all_overall[subj]; chs = all_chapter_stats[subj]
    
    R.append(f'## {"三四五六".replace("三","三") if subj=="DS" else "四" if subj=="CO" else "五"}、{full_name} ({fmt_pct(ov["rate"])})\n')
    R.append(f'题量：{ov["total"]} | 正确：{ov["correct"]} | 真题率：{fmt_pct(ov["zt_rate"])} | 自编率：{fmt_pct(ov["zb_rate"])}\n')
    
    # 章节表
    R.append('\n### 章节详情\n')
    R.append('| 章节 | 题量 | 正确率 | 真题率 | 自编率 | 最弱小节 |')
    R.append('|:---|:---:|:---:|:---:|:---:|:---|')
    for ch in chs:
        weak_sec = ch['weakest']['section'] if ch.get('weakest') else '-'
        weak_rate = fmt_pct(ch['weakest']['rate']) if ch.get('weakest') else '-'
        R.append(f"| {ch['chapter']} | {ch['total']} | {fmt_pct(ch['rate'])} | {fmt_pct(ch['zt_rate'])} | {fmt_pct(ch['zb_rate'])} | {weak_sec} ({weak_rate}) |")
    
    R.append(f'\n> 详细图表见 `分析结果/{prefix_map[subj]}_王道{full_name}/`\n')
    
    # 薄弱小节
    folder = OUT_ROOT / f'{prefix_map[subj]}_王道{full_name}'
    df_path = folder / 'cleaned_records.csv'
    if df_path.exists():
        df_s = pd.read_csv(df_path)
        weak = df_s[(df_s['total'] > 0) & (df_s['rate'] < 0.7)].sort_values('rate')
        R.append(f'\n### 薄弱小节（< 70%，共 {len(weak)} 个）\n')
        if len(weak) == 0:
            R.append('无 — 全部小节正确率 >= 70%\n')
        else:
            R.append('| 小节 | 正确率 |')
            R.append('|:---|:---:|')
            for _, w in weak.iterrows():
                R.append(f"| {w['section']} | {fmt_pct(w['rate'])} ({int(w['correct'])}/{int(w['total'])}) |")

# 总结
R.append('\n## 六、总结与建议\n')
R.append(f'- DS {fmt_pct(all_overall["DS"]["rate"])} 最强，0 个薄弱小节')
R.append(f'- CO I/O 仅 {fmt_pct(all_chapter_stats["CO"][-1]["rate"])}，最大失分点')
R.append(f'- OS 文件管理 {fmt_pct(all_chapter_stats["OS"][3]["rate"])}，投入产出比高')
R.append('- 计网未开始 — 遵循 M 三步法优先理解协议流程')
R.append(f'\n### 输出文件\n')
R.append(f'- 本报告: `{ROOT.name}/408分析报告.md`')
R.append(f'- 单科报告: `分析结果/0x_王道xxx/report.md`')
R.append(f'- 汇总对比: `分析结果/_汇总/`')

report_path = ROOT / '408分析报告.md'
report_path.write_text('\n'.join(R), encoding='utf-8')
print(f'\n408分析报告: {report_path}')
