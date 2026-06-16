"""
408 单科分析：DS / CO / OS 各生成独立报告+图表
"""
from pathlib import Path
import warnings, re
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib import font_manager
from openpyxl import load_workbook

SRC = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
OUT_DIR = Path(r'C:\Users\34406\Documents\Obsidian Vault\数据分析')
FIG = OUT_DIR / 'figures'
FIG.mkdir(parents=True, exist_ok=True)

# ============================================================
# 中文字体
# ============================================================
font_candidates = [
    'C:/Windows/Fonts/msyh.ttc', 'C:/Windows/Fonts/simhei.ttf',
    'C:/Windows/Fonts/simsun.ttc',
    '/System/Library/Fonts/PingFang.ttc',
    '/usr/share/fonts/truetype/wqy/wqy-microhei.ttc',
]
for fp in font_candidates:
    if Path(fp).exists():
        font_manager.fontManager.addfont(fp)
        plt.rcParams['font.family'] = font_manager.FontProperties(fname=fp).get_name()
        break
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun', 'WenQuanYi Micro Hei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['figure.dpi'] = 100
sns.set_style('darkgrid')
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun', 'WenQuanYi Micro Hei', 'Arial Unicode MS']
plt.rcParams['font.family'] = 'sans-serif'

# ============================================================
# 数据读取
# ============================================================
wb = load_workbook(SRC)

def safe_int(v):
    if v is None: return None
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str) and v.startswith('='): return None
    try: return int(v)
    except: return None

subjects = [
    ('DS', '王道数据结构', '数据结构'),
    ('CO', '王道计组', '计算机组成原理'),
    ('OS', '王道操作系统', '操作系统'),
]

for subj, sheet_name, full_name in subjects:
    ws = wb[sheet_name]
    heji_row = None
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).replace(' ', '') == '合计':
            heji_row = r; break
    
    records = []
    current_ch = None
    ch_data = {}
    
    for row in range(5, heji_row or ws.max_row):
        a = ws.cell(row=row, column=1).value
        if a is None: continue
        a_s = str(a).strip()
        if a_s.startswith('第') and '章' in a_s:
            current_ch = a_s
            ch_data[current_ch] = {'total':0, 'correct':0, 'zt_t':0, 'zt_c':0, 'sections':[]}
            continue
        if current_ch is None: continue
        
        c_t = safe_int(ws.cell(row=row, column=3).value)
        c_c = safe_int(ws.cell(row=row, column=4).value)
        f_zt = safe_int(ws.cell(row=row, column=6).value)
        g_ztc = safe_int(ws.cell(row=row, column=7).value)
        if c_t is None or c_t == 0: continue
        c_c = c_c or 0; f_zt = f_zt or 0; g_ztc = g_ztc or 0
        zb_t = c_t - f_zt; zb_c = c_c - g_ztc
        
        rate = c_c / c_t
        zt_rate = g_ztc / f_zt if f_zt else None
        zb_rate = zb_c / zb_t if zb_t else None
        
        records.append({
            'chapter': current_ch, 'section': a_s,
            'total': c_t, 'correct': c_c, 'rate': rate,
            'zt_t': f_zt, 'zt_c': g_ztc, 'zt_rate': zt_rate,
            'zb_t': zb_t, 'zb_c': zb_c, 'zb_rate': zb_rate,
        })
        ch_data[current_ch]['total'] += c_t
        ch_data[current_ch]['correct'] += c_c
        ch_data[current_ch]['zt_t'] += f_zt
        ch_data[current_ch]['zt_c'] += g_ztc
        ch_data[current_ch]['sections'].append(records[-1])
    
    df = pd.DataFrame(records)
    if df.empty: continue
    
    total_all = df['total'].sum()
    correct_all = df['correct'].sum()
    zt_all = df['zt_t'].sum()
    zt_c = df['zt_c'].sum()
    overall_rate = correct_all / total_all
    zt_rate = zt_c / zt_all if zt_all else None
    zb_rate = (correct_all - zt_c) / (total_all - zt_all) if (total_all - zt_all) else None
    
    ch_list = []
    for ch_name, cd in ch_data.items():
        ch_t = cd['total']; ch_c = cd['correct']
        ch_r = ch_c / ch_t
        zt_r = cd['zt_c'] / cd['zt_t'] if cd['zt_t'] else None
        zb_t = ch_t - cd['zt_t']; zb_c = ch_c - cd['zt_c']
        zb_r = zb_c / zb_t if zb_t else None
        secs = sorted(cd['sections'], key=lambda x: x['rate'])
        ch_list.append({
            'chapter': ch_name, 'chapter_no': int(re.search(r'\d+', ch_name).group()),
            'total': ch_t, 'correct': ch_c, 'rate': ch_r,
            'zt_rate': zt_r, 'zb_rate': zb_r,
            'sections': cd['sections'],
            'weakest': secs[0], 'strongest': secs[-1],
        })
    ch_list.sort(key=lambda x: x['chapter_no'])
    
    # ============================================================
    # 图表 1: 章节正确率 + 题量分布 (2in1)
    # ============================================================
    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    
    # 1a: 章节正确率横向条形图
    labels = [c['chapter'].replace('第','').replace('章','') for c in ch_list]
    rates_pct = [c['rate']*100 for c in ch_list]
    colors_ch = [plt.cm.RdYlGn(r/100) for r in rates_pct]
    bars = axes[0].barh(range(len(ch_list)), rates_pct, color=colors_ch, edgecolor='white', height=0.6)
    axes[0].set_yticks(range(len(ch_list)))
    axes[0].set_yticklabels(labels, fontsize=11)
    axes[0].set_xlim(0, 108)
    axes[0].set_xlabel('正确率 (%)')
    axes[0].set_title('各章正确率', fontsize=14, fontweight='bold')
    axes[0].axvline(x=70, color='#E74C3C', linestyle='--', alpha=0.4)
    axes[0].axvline(x=80, color='#F39C12', linestyle='--', alpha=0.4)
    axes[0].axvline(x=90, color='#27AE60', linestyle='--', alpha=0.4)
    for i, (r, ch) in enumerate(zip(rates_pct, ch_list)):
        axes[0].text(r + 1, i, f'{r:.0f}% ({ch["correct"]}/{ch["total"]})', va='center', fontsize=9)
    
    # 1b: 题量 & 正确率气泡图
    sizes = [c['total'] * 2 for c in ch_list]
    scatter = axes[1].scatter([c['total'] for c in ch_list], rates_pct, 
                              s=sizes, c=rates_pct, cmap='RdYlGn', vmin=55, vmax=100,
                              edgecolors='white', alpha=0.85)
    for i, ch in enumerate(ch_list):
        axes[1].annotate(labels[i], (ch['total'], rates_pct[i]), 
                        xytext=(5, 5), textcoords='offset points', fontsize=8)
    axes[1].set_xlabel('题量'); axes[1].set_ylabel('正确率 (%)')
    axes[1].set_title('题量 vs 正确率 (气泡大小=题量)', fontsize=14, fontweight='bold')
    axes[1].set_ylim(50, 105)
    axes[1].axhline(y=70, color='#E74C3C', linestyle='--', alpha=0.4)
    plt.colorbar(scatter, ax=axes[1], label='正确率 (%)')
    
    fig.suptitle(f'{full_name} 章节总览', fontsize=16, fontweight='bold', y=1.02)
    fig.tight_layout()
    fig.savefig(FIG / f'single_{subj}_overview.png', dpi=150, bbox_inches='tight')
    plt.close()
    
    # ============================================================
    # 图表 2: 小节排名 + 真题vs自编 (2in1)
    # ============================================================
    fig, axes = plt.subplots(1, 2, figsize=(18, 7))
    
    # 2a: 小节排名
    df_sorted = df.sort_values('rate')
    short_names = df_sorted['section'].apply(lambda s: re.sub(r'^\d+\.\d+\s*', '', s)[:20])
    colors_s = ['#E74C3C' if r < 0.7 else '#F39C12' if r < 0.8 else '#27AE60' for r in df_sorted['rate']]
    axes[0].barh(range(len(df_sorted)), df_sorted['rate']*100, color=colors_s, edgecolor='white', height=0.7)
    axes[0].set_yticks(range(len(df_sorted)))
    axes[0].set_yticklabels(short_names, fontsize=7)
    axes[0].set_xlim(0, 108)
    axes[0].axvline(x=70, color='#E74C3C', linestyle='--', alpha=0.5)
    axes[0].axvline(x=80, color='#F39C12', linestyle='--', alpha=0.5)
    axes[0].axvline(x=90, color='#27AE60', linestyle='--', alpha=0.5)
    axes[0].set_title('小节正确率排名', fontsize=14, fontweight='bold')
    for i, (_, r) in enumerate(df_sorted.head(3).iterrows()):
        axes[0].text(r['rate']*100+1, i, f"{int(r['correct'])}/{int(r['total'])}", va='center', fontsize=7)
    
    # 2b: 真题 vs 自编
    labels2 = [c['chapter'].replace('第','').replace('章','') for c in ch_list]
    x = np.arange(len(labels2)); w = 0.3
    zt_rates_ch = [(c['zt_rate']*100 if c['zt_rate'] else 0) for c in ch_list]
    zb_rates_ch = [(c['zb_rate']*100 if c['zb_rate'] else 0) for c in ch_list]
    all_rates_ch = [c['rate']*100 for c in ch_list]
    
    axes[1].bar(x - w, zt_rates_ch, w, label='真题', color='#2E86AB', edgecolor='white')
    axes[1].bar(x, zb_rates_ch, w, label='自编', color='#F18F01', edgecolor='white')
    axes[1].bar(x + w, all_rates_ch, w, label='总体', color='#27AE60', edgecolor='white', alpha=0.55)
    axes[1].set_xticks(x); axes[1].set_xticklabels(labels2, fontsize=9, rotation=30)
    axes[1].set_ylabel('正确率 (%)'); axes[1].set_ylim(0, 112)
    axes[1].set_title('真题 vs 自编 各章对比', fontsize=14, fontweight='bold')
    axes[1].legend(fontsize=10, loc='lower left')
    
    fig.suptitle(f'{full_name} 小节详情 & 真题/自编拆分', fontsize=16, fontweight='bold', y=1.02)
    fig.tight_layout()
    fig.savefig(FIG / f'single_{subj}_detail.png', dpi=150, bbox_inches='tight')
    plt.close()
    
    # ============================================================
    # 图表 3: 正确率分布直方图
    # ============================================================
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.hist(df['rate']*100, bins=12, color='#2E86AB', edgecolor='white', alpha=0.8)
    ax.axvline(x=overall_rate*100, color='#E74C3C', linestyle='-', linewidth=2.5, label=f'平均 {overall_rate*100:.1f}%')
    ax.axvline(x=70, color='#E74C3C', linestyle='--', alpha=0.5)
    ax.axvline(x=80, color='#F39C12', linestyle='--', alpha=0.5)
    ax.axvline(x=90, color='#27AE60', linestyle='--', alpha=0.5)
    ax.set_xlabel('正确率 (%)'); ax.set_ylabel('小节数')
    ax.set_title(f'{full_name} 小节正确率分布', fontsize=14, fontweight='bold')
    ax.legend(fontsize=11)
    
    # Add count labels on bars
    for patch in ax.patches:
        h = patch.get_height()
        if h > 0:
            ax.text(patch.get_x()+patch.get_width()/2, h+0.3, str(int(h)), ha='center', fontsize=10)
    
    fig.tight_layout()
    fig.savefig(FIG / f'single_{subj}_dist.png', dpi=150, bbox_inches='tight')
    plt.close()
    
    # ============================================================
    # 生成 Markdown 报告
    # ============================================================
    def fmt_pct(v):
        if v is None: return '-'
        return f'{v*100:.1f}%'
    
    lines = []
    lines.append(f'# {full_name} — 一刷完成分析\n')
    lines.append(f'> 总题量：{total_all} | 正确：{correct_all} | 正确率：{fmt_pct(overall_rate)}\n')
    
    if zt_rate is not None:
        lines.append(f'**真题**：{zt_c}/{zt_all} = {fmt_pct(zt_rate)} | **自编**：{correct_all-zt_c}/{total_all-zt_all} = {fmt_pct(zb_rate)}\n')
    
    lines.append(f'![章节总览](figures/single_{subj}_overview.png)\n')
    
    # 章节表
    lines.append('## 章节详情\n')
    lines.append('| 章节 | 题量 | 正确率 | 真题率 | 自编率 | 最弱小节 | 最强小节 |')
    lines.append('|:---|:---:|:---:|:---:|:---:|:---|:---|')
    for ch in ch_list:
        lines.append(f"| {ch['chapter']} | {ch['total']} | {fmt_pct(ch['rate'])} | {fmt_pct(ch['zt_rate'])} | {fmt_pct(ch['zb_rate'])} | {ch['weakest']['section']} ({fmt_pct(ch['weakest']['rate'])}) | {ch['strongest']['section']} ({fmt_pct(ch['strongest']['rate'])}) |")
    
    lines.append(f'\n![小节详情](figures/single_{subj}_detail.png)\n')
    
    # 薄弱小节
    weak = df[df['rate'] < 0.7].sort_values('rate')
    lines.append(f'## 薄弱小节（正确率 < 70%，共 {len(weak)} 个）\n')
    if len(weak) == 0:
        lines.append('**无** — 全部小节正确率 >= 70%\n')
    else:
        lines.append('| 小节 | 正确率 | 真题率 | 自编率 | 诊断 |')
        lines.append('|:---|:---:|:---:|:---:|:---|')
        for _, w in weak.iterrows():
            diag = ''
            if w['zt_rate'] is not None and w['zb_rate'] is not None:
                if w['zt_rate'] > 0.8 and (w['zb_rate'] or 0) < 0.6:
                    diag = '真题强/自编弱 → 自编题偏难'
                elif (w['zt_rate'] or 0) < 0.6 and w['zb_rate'] > 0.7:
                    diag = '自编强/真题弱 → 真题思路未掌握'
                else:
                    diag = '两边都弱 → 需重点突破'
            lines.append(f"| {w['section']} | {fmt_pct(w['rate'])} ({int(w['correct'])}/{int(w['total'])}) | {fmt_pct(w['zt_rate'])} | {fmt_pct(w['zb_rate'])} | {diag} |")
    
    # 掌握牢固
    strong = df[df['rate'] >= 0.9].sort_values('rate', ascending=False)
    lines.append(f'\n## 掌握牢固（正确率 >= 90%，共 {len(strong)} 个）\n')
    if len(strong) > 0:
        lines.append('| 小节 | 正确率 |')
        lines.append('|:---|:---:|')
        for _, s in strong.iterrows():
            lines.append(f"| {s['section']} | {fmt_pct(s['rate'])} ({int(s['correct'])}/{int(s['total'])}) |")
    
    lines.append(f'\n![分布](figures/single_{subj}_dist.png)\n')
    
    # 生成图表清单
    lines.append('## 生成图表\n')
    for f in [f'single_{subj}_overview.png', f'single_{subj}_detail.png', f'single_{subj}_dist.png']:
        lines.append(f'- `figures/{f}`')
    
    report_path = OUT_DIR / f'{full_name}分析报告.md'
    report_path.write_text('\n'.join(lines), encoding='utf-8')
    print(f'[{subj}] {full_name} 分析报告.md  ({len(ch_list)}章, {len(df)}小节, {fmt_pct(overall_rate)})')

print('\n全部单科报告生成完成')
