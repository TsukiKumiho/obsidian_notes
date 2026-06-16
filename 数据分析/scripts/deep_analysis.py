"""
408 深度分析可视化 — 5 个新维度
"""
from pathlib import Path
import warnings, re
warnings.filterwarnings("ignore")
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib import font_manager
from openpyxl import load_workbook

SRC = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
FIG = Path(r'C:\Users\34406\Documents\Obsidian Vault\数据分析\figures')
FIG.mkdir(parents=True, exist_ok=True)

# === 字体 ===
for fp in ['C:/Windows/Fonts/msyh.ttc','C:/Windows/Fonts/simhei.ttf','C:/Windows/Fonts/simsun.ttc']:
    if Path(fp).exists():
        font_manager.fontManager.addfont(fp)
        plt.rcParams['font.family'] = font_manager.FontProperties(fname=fp).get_name()
        break
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False; plt.rcParams['figure.dpi'] = 100
sns.set_style('darkgrid')
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'SimSun', 'Arial Unicode MS']
plt.rcParams['font.family'] = 'sans-serif'

# === 数据 ===
wb = load_workbook(SRC)
def si(v):
    if v is None: return None
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str) and v.startswith('='): return None
    try: return int(v)
    except: return None

sheets = {'DS': '王道数据结构', 'CO': '王道计组', 'OS': '王道操作系统'}
names = {'DS': '数据结构', 'CO': '计组', 'OS': '操作系统'}
all_data = []

for subj, sn in sheets.items():
    ws = wb[sn]
    heji = None
    for r in range(1, ws.max_row+1):
        if ws.cell(row=r, column=1).value and str(ws.cell(row=r, column=1).value).replace(' ','') == '合计':
            heji = r; break
    cur = None
    for row in range(5, heji):
        a = ws.cell(row=row, column=1).value
        if a is None: continue
        a_s = str(a).strip()
        if a_s.startswith('第') and '章' in a_s:
            cur = a_s; continue
        c_t = si(ws.cell(row=row, column=3).value); c_c = si(ws.cell(row=row, column=4).value)
        f_zt = si(ws.cell(row=row, column=6).value); g_ztc = si(ws.cell(row=row, column=7).value)
        if c_t and c_t > 0:
            c_c = c_c or 0; f_zt = f_zt or 0; g_ztc = g_ztc or 0
            all_data.append({
                'subj': subj, 'subj_name': names[subj], 'chapter': cur,
                'section': a_s, 'total': c_t, 'correct': c_c,
                'wrong': c_t - c_c, 'rate': c_c / c_t,
                'zt_t': f_zt, 'zt_zt': g_ztc, 'zb_t': c_t - f_zt, 'zb_c': c_c - g_ztc,
            })

df = pd.DataFrame(all_data)

# Chapter-level aggregation per subject
ch_agg = df.groupby(['subj', 'subj_name', 'chapter']).agg(
    total=('total','sum'), correct=('correct','sum'), wrong=('wrong','sum'),
    zt_t=('zt_t','sum'), zt_c=('zt_zt','sum'), zb_t=('zb_t','sum'), zb_c=('zb_c','sum'),
    sections=('section','count')
).reset_index()
ch_agg['rate'] = ch_agg['correct'] / ch_agg['total']
ch_agg['zt_rate'] = (ch_agg['zt_c'] / ch_agg['zt_t'].replace(0, np.nan)) * 100
ch_agg['zb_rate'] = (ch_agg['zb_c'] / ch_agg['zb_t'].replace(0, np.nan)) * 100
ch_agg['gap'] = ch_agg['zt_rate'] - ch_agg['zb_rate']

def save(fig, name):
    fig.savefig(FIG / name, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f'  {name}')

# ============================================================
# 图 1: 错题绝对量 + 正确率双轴图
# ============================================================
fig, axes = plt.subplots(1, 3, figsize=(20, 7))
for ax_i, subj in enumerate(['DS', 'CO', 'OS']):
    ax = axes[ax_i]
    d = ch_agg[ch_agg['subj'] == subj].sort_values('rate')
    labels = d['chapter'].str.replace('第','').str.replace('章','')
    x = np.arange(len(d)); w = 0.35
    bars = ax.bar(x - w/2, d['wrong'], w, label='错题数', color='#E74C3C', edgecolor='white', alpha=0.85)
    ax2 = ax.twinx()
    ax2.plot(x + w/2, d['rate']*100, 'D-', color='#2E86AB', linewidth=2.5, markersize=10, label='正确率')
    ax.set_xticks(x); ax.set_xticklabels(labels, fontsize=8, rotation=30)
    ax.set_ylabel('错题数', color='#E74C3C'); ax2.set_ylabel('正确率 (%)', color='#2E86AB')
    ax.set_title(names[subj], fontsize=14, fontweight='bold')
    for i, (wv, r) in enumerate(zip(d['wrong'], d['rate'])):
        ax.text(i - w/2, wv + 0.5, str(int(wv)), ha='center', fontsize=8, fontweight='bold')
    lines1, labels1 = ax.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax.legend(lines1+lines2, labels1+labels2, loc='upper left', fontsize=9)
fig.suptitle('错题绝对量 vs 正确率（按正确率升序）', fontsize=16, fontweight='bold', y=1.02)
fig.tight_layout()
save(fig, '408_wrong_count.png')

# ============================================================
# 图 2: 题量-正确率象限图
# ============================================================
fig, axes = plt.subplots(1, 3, figsize=(20, 6.5))
for ax_i, subj in enumerate(['DS', 'CO', 'OS']):
    ax = axes[ax_i]
    d = ch_agg[ch_agg['subj'] == subj]
    labels = d['chapter'].str.replace('第','').str.replace('章','')
    sizes = d['total'] * 2
    scatter = ax.scatter(d['total'], d['rate']*100, s=sizes, c=d['rate']*100,
                         cmap='RdYlGn', vmin=55, vmax=100, edgecolors='white', alpha=0.85)
    # Quadrant lines
    med_total = d['total'].median(); med_rate = d['rate'].median() * 100
    ax.axhline(y=med_rate, color='#999', linestyle='--', alpha=0.5)
    ax.axvline(x=med_total, color='#999', linestyle='--', alpha=0.5)
    # Labels
    for idx, (_, r) in enumerate(d.iterrows()):
        lbl = labels.iloc[idx]
        quadrant = ''
        if r['total'] > med_total and r['rate']*100 < med_rate: quadrant = ' !!!'  # high vol, low rate
        ax.annotate(lbl + quadrant, (r['total'], r['rate']*100), xytext=(5,5), textcoords='offset points', fontsize=8)
    ax.set_xlabel('题量'); ax.set_ylabel('正确率 (%)')
    ax.set_title(f'{names[subj]} (中位数: {med_total:.0f}题 / {med_rate:.0f}%)', fontsize=13, fontweight='bold')
    plt.colorbar(scatter, ax=ax, label='正确率 (%)', shrink=0.8)
fig.suptitle('题量 vs 正确率 象限图（右上=大量+高分, 左下=少量+低分）', fontsize=16, fontweight='bold', y=1.02)
fig.tight_layout()
save(fig, '408_quadrant.png')

# ============================================================
# 图 3: 跨科关联 (CO vs OS 重叠知识点)
# ============================================================
# CO ch1(概述) vs OS ch1(概述), CO ch7(I/O) vs OS ch5(I/O), CO ch3(存储) vs OS ch3(内存)
overlap = [
    ('概述', 'CO', '第1章 计算机系统概述', 'OS', '第1章 计算机系统概述'),
    ('I/O', 'CO', '第7章 输入/输出系统', 'OS', '第5章 输入/输出管理'),
    ('存储/内存', 'CO', '第3章 存储系统', 'OS', '第3章 内存管理'),
]

fig, ax = plt.subplots(figsize=(10, 6))
x = np.arange(len(overlap)); w = 0.3
for i, (label, s1, ch1, s2, ch2) in enumerate(overlap):
    r1 = ch_agg[(ch_agg['subj']==s1)&(ch_agg['chapter']==ch1)]['rate'].values
    r2 = ch_agg[(ch_agg['subj']==s2)&(ch_agg['chapter']==ch2)]['rate'].values
    r1v = r1[0]*100 if len(r1) else 0
    r2v = r2[0]*100 if len(r2) else 0
    ax.bar(i - w/2, r1v, w, label='计组' if i==0 else '', color='#2E86AB', edgecolor='white')
    ax.bar(i + w/2, r2v, w, label='操作系统' if i==0 else '', color='#F18F01', edgecolor='white')
    ax.text(i - w/2, r1v + 1, f'{r1v:.0f}%', ha='center', fontsize=10, fontweight='bold')
    ax.text(i + w/2, r2v + 1, f'{r2v:.0f}%', ha='center', fontsize=10, fontweight='bold')
    diff = r1v - r2v
    ax.annotate(f'差{diff:+.0f}%', (i, max(r1v, r2v) + 8), ha='center', fontsize=11, fontweight='bold',
                color='#E74C3C' if abs(diff) > 10 else '#666')
ax.set_xticks(x); ax.set_xticklabels([o[0] for o in overlap], fontsize=13)
ax.set_ylabel('正确率 (%)'); ax.set_ylim(0, 100)
ax.legend(fontsize=12); ax.set_title('CO vs OS 重叠知识点正确率对比', fontsize=15, fontweight='bold')
fig.tight_layout()
save(fig, '408_cross_subject.png')

# ============================================================
# 图 4: 目标差距量化 (目标 83% = 125分/150)
# ============================================================
target = 83
fig, axes = plt.subplots(1, 3, figsize=(20, 7))
for ax_i, subj in enumerate(['DS', 'CO', 'OS']):
    ax = axes[ax_i]
    d = ch_agg[ch_agg['subj'] == subj].sort_values('rate')
    labels = d['chapter'].str.replace('第','').str.replace('章','')
    current = d['rate'] * 100
    gap = target - current
    above = [max(0, t - target) for t in current]
    below = [max(0, target - t) for t in current]
    
    ax.barh(range(len(d)), current, color=['#27AE60' if r >= target else '#E74C3C' for r in current],
            edgecolor='white', height=0.6)
    ax.axvline(x=target, color='#2E86AB', linestyle='-', linewidth=2, label=f'目标 {target}%')
    ax.set_yticks(range(len(d))); ax.set_yticklabels(labels, fontsize=10)
    ax.set_xlim(0, 105); ax.set_xlabel('正确率 (%)')
    ax.set_title(f'{names[subj]} (目标 {target}%)', fontsize=14, fontweight='bold')
    ax.legend(fontsize=10)
    for i, (r, g) in enumerate(zip(current, gap)):
        if g > 0:
            ax.text(r + 1, i, f'-{g:.0f}pp', va='center', fontsize=9, color='#E74C3C', fontweight='bold')
        else:
            ax.text(r + 1, i, '✓', va='center', fontsize=10, color='#27AE60', fontweight='bold')
fig.suptitle('各章距 125 分目标线（83%）的差距', fontsize=16, fontweight='bold', y=1.02)
fig.tight_layout()
save(fig, '408_target_gap.png')

# ============================================================
# 图 5: 真题-自编差距显著性
# ============================================================
fig, axes = plt.subplots(1, 3, figsize=(20, 7))
for ax_i, subj in enumerate(['DS', 'CO', 'OS']):
    ax = axes[ax_i]
    d = ch_agg[ch_agg['subj'] == subj].sort_values('gap')
    labels = d['chapter'].str.replace('第','').str.replace('章','')
    gaps = d['gap'].values
    colors_g = ['#E74C3C' if abs(g) > 15 else '#F39C12' if abs(g) > 8 else '#27AE60' for g in gaps]
    ax.barh(range(len(d)), gaps, color=colors_g, edgecolor='white', height=0.6)
    ax.set_yticks(range(len(d))); ax.set_yticklabels(labels, fontsize=10)
    ax.axvline(x=0, color='black', linewidth=1)
    ax.axvline(x=8, color='#F39C12', linestyle='--', alpha=0.5)
    ax.axvline(x=-8, color='#F39C12', linestyle='--', alpha=0.5)
    ax.set_xlabel('真题率 - 自编率 (pp)')
    ax.set_title(f'{names[subj]}', fontsize=14, fontweight='bold')
    for i, g in enumerate(gaps):
        ax.text(g + (1 if g >= 0 else -8), i, f'{g:+.0f}pp', va='center', fontsize=9, fontweight='bold')
fig.suptitle('真题 vs 自编 差距显著性（|gap|>15pp 显著, >8pp 值得关注）', fontsize=16, fontweight='bold', y=1.02)
fig.tight_layout()
save(fig, '408_gap_significance.png')

print(f'\n5 charts saved to {FIG}')
