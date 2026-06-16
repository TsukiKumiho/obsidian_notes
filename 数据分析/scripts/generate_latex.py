# -*- coding: utf-8 -*-
"""生成 408分析报告.tex — 从 CSV 数据动态填充表格"""
from pathlib import Path
import pandas as pd, numpy as np, re
from openpyxl import load_workbook

ROOT = Path(r'C:\Users\34406\Documents\Obsidian Vault\数据分析')
RESULT = ROOT / '分析结果'
FIG = ROOT / 'figures'
OUT = ROOT / 'latex'
OUT.mkdir(parents=True, exist_ok=True)

def fmt_pct(v, raw=False):
    if v is None or pd.isna(v): return '—'
    return f'{v*100:.1f}' if raw else f'{v*100:.1f}\\%'

def rate_cell(v):
    """Return LaTeX cell with color — only for 真题/自编 tables"""
    if v is None or pd.isna(v): return '—'
    p = v * 100
    if p >= 90: color = 'greenrate'
    elif p >= 80: color = 'yellowrate'
    elif p >= 70: return f'{p:.1f}\\%'
    else: color = 'redrate'
    return f'\\cellcolor{{{color}}}{{{p:.1f}\\%}}'

def rate_plain(v):
    """Plain rate without cellcolor — for chapter detail tables"""
    if v is None or pd.isna(v): return '—'
    return f'{v*100:.1f}\\%'

# ============================================================
# Load data
# ============================================================
folders = {'DS': '04_王道数据结构', 'CO': '03_王道计算机组成原理', 'OS': '02_王道操作系统'}
names = {'DS': '数据结构', 'CO': '计算机组成原理', 'OS': '操作系统'}
data = {}
for subj, folder in folders.items():
    data[subj] = {
        'chapter': pd.read_csv(RESULT / folder / 'chapter_summary.csv'),
        'records': pd.read_csv(RESULT / folder / 'cleaned_records.csv'),
        'overall': pd.read_csv(RESULT / folder / 'overall_summary.csv').iloc[0],
        'folder': folder,
    }

total_t = sum(int(data[s]['overall']['first_total']) for s in ['DS','CO','OS'])
total_c = sum(int(data[s]['overall']['first_correct']) for s in ['DS','CO','OS'])
overall_rate = total_c / total_t
est_score = round(overall_rate * 150)

# ============================================================
# Build LaTeX
# ============================================================
L = []

def add(*lines):
    for line in lines:
        if line == '':
            L.append('')
        else:
            L.append(line)

# --- Preamble ---
add(
    r'\documentclass[12pt,a4paper]{ctexart}',
    '',
    r'\usepackage[top=2.5cm,bottom=2.5cm,left=2cm,right=2cm]{geometry}',
    r'\usepackage{graphicx}',
    r'\usepackage[table]{xcolor}',
    r'\usepackage{booktabs}',
    r'\usepackage{float}',
    r'\usepackage{hyperref}',
    r'\usepackage{tabularx}',
    r'\usepackage{array}',
    r'\usepackage{caption}',
    r'\usepackage{enumitem}',
    r'\usepackage{zhnumber}',
    '',
    r'\definecolor{greenrate}{HTML}{C6EFCE}',
    r'\definecolor{yellowrate}{HTML}{FFEB9C}',
    r'\definecolor{redrate}{HTML}{FFC7CE}',
    r'\definecolor{headerbg}{HTML}{2F5496}',
    r'\definecolor{headerfg}{HTML}{FFFFFF}',
    r'\definecolor{grayrow}{HTML}{F2F2F2}',
    r'\definecolor{yellowrow}{HTML}{FFF2CC}',
    r'\definecolor{pinkrow}{HTML}{F4B4C2}',
    '',
    r'\graphicspath{{../figures/}{../分析结果/}}',
    '',
    r'\hypersetup{colorlinks=true,linkcolor=blue,urlcolor=blue}',
    r'\setenumerate[1]{itemsep=0pt,partopsep=0pt,parsep=0pt,topsep=0pt}',
    r'\setitemize[1]{itemsep=0pt,partopsep=0pt,parsep=0pt,topsep=0pt}',
    r'\setlength{\tabcolsep}{4pt}',  # compact table padding
    '',
    r'\title{\textbf{408 计算机统考 — 全面分析报告}}',
    r'\author{数据管道: 习题册.xlsx → export\_xlsx\_to\_csv.py → analyze\_csv\_folder.py}',
    r'\date{\today}',
    '',
    r'\begin{document}',
    r'\maketitle',
    r'\thispagestyle{empty}',
    r'\newpage',
    r'\tableofcontents',
    r'\newpage',
    '',
)

# --- Helper to write tables ---
def begin_tabular(cols, fontsize=r'\footnotesize', resize=False):
    add(f'\\begin{{table}}[H]', f'{fontsize}', f'\\centering')
    if resize:
        add(r'\resizebox{\textwidth}{!}{')
    add(f'\\begin{{tabular}}{{{cols}}}', r'\hline')

def end_tabular(caption='', label='', resize=False):
    add(r'\hline', r'\end{tabular}')
    if resize:
        add('}')
    if caption:
        add(f'\\caption{{{caption}}}')
    if label:
        add(f'\\label{{{label}}}')
    add(r'\end{table}', '')

# ============================================================
# Section 1: 总览
# ============================================================
add(r'\section{四科总览}', '')

begin_tabular('lrrrrrr')
add(r'\rowcolor{headerbg}\textcolor{headerfg}{\textbf{科目}} & \textcolor{headerfg}{\textbf{题量}} & \textcolor{headerfg}{\textbf{正确数}} & \textcolor{headerfg}{\textbf{正确率}} & \textcolor{headerfg}{\textbf{已做章节}} & \textcolor{headerfg}{\textbf{已做小节}} & \textcolor{headerfg}{\textbf{状态}} \\')
for subj in ['DS', 'CO', 'OS']:
    d = data[subj]; ov = d['overall']
    t = int(ov['first_total']); c = int(ov['first_correct'])
    ch_n = len(d['chapter']); sec_n = int(ov['sections_with_first_data'])
    add(f'\\textbf{{{names[subj]}}} & {t} & {c} & {rate_cell(ov["first_rate"])} & {ch_n} & {sec_n} & 一刷完成 \\\\')
add(r'\rowcolor{grayrow} 计算机网络 & — & — & 未开始 & — & — & 待开始 \\\\')
end_tabular('四科完成总览', 'tab:overview')

add(f'\\textbf{{三科合计}}：{total_c}/{total_t} = \\textbf{{{overall_rate*100:.1f}\\%}}，408 估分约 \\textbf{{{est_score} 分}}（满分 150）。')
add('估分方式：章节正确率 $\\times$ 150 直接折算。实际真题难度通常高于章节练习，得分可能略低 5—10 分。')
add('')

# ============================================================
# Section 2: 真题 vs 自编
# ============================================================
wb_xl = load_workbook(r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx')
sheets_xl = {'DS': '王道数据结构', 'CO': '王道计组', 'OS': '王道操作系统'}
def si(v):
    if v is None: return None
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str) and v.startswith('='): return None
    try: return int(v)
    except: return None

add(r'\section{真题 vs 自编 章节对比}', '')
add(r'\begin{figure}[H]', r'\centering',
    r'\includegraphics[width=0.92\textwidth]{408_zhenti_vs_zibian.png}',
    r'\caption{真题 vs 自编 vs 总体 各章正确率对比}', r'\end{figure}', '')

for subj in ['DS', 'CO', 'OS']:
    ws = wb_xl[sheets_xl[subj]]
    heji = None
    for r in range(1, ws.max_row+1):
        if ws.cell(row=r, column=1).value and str(ws.cell(row=r, column=1).value).replace(' ','') == '合计':
            heji = r; break
    
    ch_zt = {}; cur = None
    for row in range(5, heji):
        a = ws.cell(row=row, column=1).value
        if a is None:
            continue
        a_s = str(a).strip()
        if a_s.startswith('第') and '章' in a_s:
            cur = a_s; ch_zt[cur] = {'zt_t':0,'zt_c':0,'zb_t':0,'zb_c':0}; continue
        c_t = si(ws.cell(row=row, column=3).value); c_c = si(ws.cell(row=row, column=4).value)
        f_zt = si(ws.cell(row=row, column=6).value); g_ztc = si(ws.cell(row=row, column=7).value)
        if c_t and c_t > 0:
            c_c = c_c or 0; f_zt = f_zt or 0; g_ztc = g_ztc or 0
            ch_zt[cur]['zt_t'] += f_zt; ch_zt[cur]['zt_c'] += g_ztc
            ch_zt[cur]['zb_t'] += c_t - f_zt; ch_zt[cur]['zb_c'] += c_c - g_ztc
    
    add(f'\\subsection{{{names[subj]}}}', '')
    begin_tabular('lrrrl')
    add(r'\rowcolor{headerbg}\textcolor{headerfg}{\textbf{章节}} & \textcolor{headerfg}{\textbf{真题正确率}} & \textcolor{headerfg}{\textbf{自编正确率}} & \textcolor{headerfg}{\textbf{差距}} & \textcolor{headerfg}{\textbf{诊断}} \\')
    for cn, cz in ch_zt.items():
        zt_r = cz['zt_c']/cz['zt_t']*100 if cz['zt_t'] else 0
        zb_r = cz['zb_c']/cz['zb_t']*100 if cz['zb_t'] else 0
        gap = zt_r - zb_r
        diag = '均衡' if abs(gap)<5 else ('真题优势' if gap>0 else '自编优势')
        cn_spaced = re.sub(r'第(\d+)章', r'第 \1 章 ', cn)
        add(f'{cn_spaced} & {rate_cell(cz["zt_c"]/cz["zt_t"]) if cz["zt_t"] else "—"} & {rate_cell(cz["zb_c"]/cz["zb_t"]) if cz["zb_t"] else "—"} & {gap:+.0f}pp & {diag} \\\\')
    end_tabular(f'{names[subj]} 真题 vs 自编')

# ============================================================
# Section 3~5: Per-subject detail
# ============================================================
sec_num = [3, 4, 5]
for i, subj in enumerate(['DS', 'CO', 'OS']):
    d = data[subj]; ov = d['overall']; ch_df = d['chapter']; cr = d['records']
    t = int(ov['first_total']); c = int(ov['first_correct']); r = ov['first_rate']
    
    add(f'\\section{{{names[subj]}（{fmt_pct(r)}）}}', '')
    add(f'题量：{t} \\quad 正确：{c} \\quad 章节：{len(ch_df)} \\quad 小节：{int(ov["sections_with_first_data"])}', '')
    
    # Chapter analysis figure
    add(r'\begin{figure}[H]', r'\centering',
        f'\\includegraphics[width=0.92\\textwidth]{{{d["folder"]}/chapter_analysis.png}}',
        r'\caption{章节与分布分析}', r'\end{figure}', '')
    
    add(f'\\subsection{{章节详情}}', '')
    add(r'\begin{itemize}')
    for _, ch in ch_df.iterrows():
        if ch['first_total'] == 0: continue
        ch_name = re.sub(r'第(\d+)章', r'第 \1 章 ', ch['chapter'])
        wn = ch.get('weakest_section', '')
        wr = ch.get('weakest_rate', np.nan)
        wd = f'最弱：{wn}（{wr:.1%}）' if (pd.notna(wr) and wr > 0 and isinstance(wn, str)) else ''
        add(f'\\item \\textbf{{{ch_name}}} — {int(ch["first_total"])} 题，正确率 {rate_plain(ch["first_rate"])}。{wd}')
    add(r'\end{itemize}', '')
    add(r'\begin{figure}[H]', r'\centering',
        f'\\includegraphics[width=0.92\\textwidth]{{{d["folder"]}/section_heatmap.png}}',
        r'\caption{小节正确率热力图}', r'\end{figure}', '')
    
    # Weak sections
    weak = cr[(cr['total'] > 0) & (cr['rate'] < 0.7)].sort_values('rate')
    add(f'\\subsection{{薄弱小节（< 70\\%，共 {len(weak)} 个）}}', '')
    if len(weak) == 0:
        add('\\textbf{无} — 全部小节正确率 $\\ge$ 70\\%。', '')
    else:
        begin_tabular('lll')
        add(r'\rowcolor{headerbg}\textcolor{headerfg}{\textbf{小节}} & \textcolor{headerfg}{\textbf{正确率}} & \textcolor{headerfg}{\textbf{建议}} \\')
        for _, w in weak.iterrows():
            rv = w['rate']*100
            tip = '\\textbf{优先二刷}' if rv < 55 else ('重点回顾' if rv < 65 else '查漏补缺')
            add(f'{w["section"]} & {rate_cell(w["rate"])} ({int(w["correct"])}/{int(w["total"])}) & {tip} \\\\')
        end_tabular(f'{names[subj]} 薄弱小节')
    
    # Strong sections
    strong = cr[(cr['total'] > 0) & (cr['rate'] >= 0.9)].sort_values('rate', ascending=False)
    if len(strong) > 0:
        add(f'\\subsection{{掌握牢固（$\\ge$ 90\\%，共 {len(strong)} 个）}}', '')
        items = [f'{s["section"]} ({int(s["correct"])}/{int(s["total"])})' for _, s in strong.head(8).iterrows()]
        add('、'.join(items) + '。', '')

# ============================================================
# Section 6: Deep Analysis
# ============================================================
add(r'\section{深度分析}', '')

deep_sections = [
    ('错题绝对量 vs 正确率', '408_wrong_count.png',
     '正确率相同时，错题量决定优先级。例如 CO I/O 方式错 22 题（55 题中），比中断错 6 题更需要时间投入。'),
    ('题量-正确率象限图', '408_quadrant.png',
     '高题量+低正确率（气泡大且偏下）的章节是最大风险区。'),
    ('跨科关联：CO vs OS 重叠知识点', '408_cross_subject.png',
     '408 命题常跨科综合。I/O 知识点 CO 62\\% vs OS 74\\%，差 12pp，问题在计组硬件层。'),
    ('目标差距量化（408 目标 125 分 → 每科 83\\%）', '408_target_gap.png',
     f'DS 85.4\\% 已达标，CO 需 +9.2pp（约 54 题），OS 需 +8.5pp（约 55 题）。'),
    ('真题-自编差距显著性', '408_gap_significance.png',
     '差距 $>$15pp 表示真题和自编考察角度差异大，需分别训练。DS 第1章(30pp)、第2章(31pp)自编远强于真题。'),
]

for title, img, desc in deep_sections:
    add(f'\\subsection{{{title}}}', '')
    add(desc, '')
    add(r'\begin{figure}[H]', r'\centering',
        f'\\includegraphics[width=0.92\\textwidth]{{{img}}}',
        f'\\caption{{{title}}}', r'\end{figure}', '')

# ============================================================
# Section 7: Summary
# ============================================================
add(r'\section{总结与建议}', '')

add(r'\subsection{当前状态}', '')
begin_tabular('ll')
add(r'\rowcolor{headerbg}\textcolor{headerfg}{\textbf{指标}} & \textcolor{headerfg}{\textbf{数值}} \\')
add(f'三科合计 & {total_c}/{total_t} = {overall_rate*100:.1f}\\% \\\\')
add(f'408 估分 & {est_score} / 150 \\\\')
add(f'已完成 & 3/4 科（DS/CO/OS） \\\\')
add(f'DS 薄弱小节 & 0 个 \\\\')
n_co = len(data['CO']['records'][(data['CO']['records']['total']>0)&(data['CO']['records']['rate']<0.7)])
n_os = len(data['OS']['records'][(data['OS']['records']['total']>0)&(data['OS']['records']['rate']<0.7)])
add(f'CO 薄弱小节 & {n_co} 个 \\\\')
add(f'OS 薄弱小节 & {n_os} 个 \\\\')
end_tabular('当前状态概览')

add(r'\subsection{优势}', '')
add(r'\begin{itemize}',
    f'\\item \\textbf{{DS {fmt_pct(data["DS"]["overall"]["first_rate"])}}}：树/图/查找/排序均 85\\%+，0 个薄弱小节',
    r'\item CO 指令系统(78\%)和总线(80\%)达到良好水平',
    r'\item OS 进程管理(79\%)题量最大但正确率稳定',
    r'\end{itemize}', '')

add(r'\subsection{短板与对策}', '')
begin_tabular('llll')
add(r'\rowcolor{headerbg}\textcolor{headerfg}{\textbf{短板}} & \textcolor{headerfg}{\textbf{当前}} & \textcolor{headerfg}{\textbf{目标}} & \textcolor{headerfg}{\textbf{对策}} \\')
add(r'CO I/O 系统 & 61.8\% & 75\% & 中断+DMA+接口，3 小节全 < 70\% \\')
add(r'CO 存储系统 & 72.3\% & 80\% & Cache+虚拟存储是 408 大题核心 \\')
add(r'OS 文件管理 & 66.7\% & 78\% & 目录结构+文件分配，70 题 \\')
add(r'计网 & 未开始 & 70\% & M 三步法：理解→分层→复盘 \\')
end_tabular('短板与对策')

add(r'\subsection{下一步行动}', '')
add(r'\begin{enumerate}',
    r'\item \textbf{计网开课} — 最后一块拼图，概念密集型',
    r'\item \textbf{CO I/O 二刷} — 仅 3 小节 76 题，2—3 天可完成',
    r'\item \textbf{OS 文件管理} — 70 题，与 I/O 穿插进行',
    r'\item \textbf{DS 维持} — 隔天 5—10 题保持手感',
    r'\end{enumerate}', '')

# --- End ---
add(r'\end{document}')

# Write .tex
tex_path = OUT / '408分析报告.tex'
tex_path.write_text('\n'.join(L), encoding='utf-8')
print(f'Generated: {tex_path} ({len(L)} lines)')
