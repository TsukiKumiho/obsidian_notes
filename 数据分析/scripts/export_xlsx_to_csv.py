from __future__ import annotations

import argparse
import ast
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_WORKBOOK = ROOT / "习题册excel.xlsx"
DEFAULT_OUT_DIR = Path(__file__).resolve().parents[1] / "csv"


def safe_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", name).strip()
    return cleaned or "sheet"


def is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.startswith("=")


def split_args(text: str) -> list[str]:
    args: list[str] = []
    current: list[str] = []
    depth = 0
    in_string = False
    i = 0
    while i < len(text):
        char = text[i]
        if char == '"':
            in_string = not in_string
            current.append(char)
        elif not in_string and char == "(":
            depth += 1
            current.append(char)
        elif not in_string and char == ")":
            depth -= 1
            current.append(char)
        elif not in_string and char == "," and depth == 0:
            args.append("".join(current).strip())
            current = []
        else:
            current.append(char)
        i += 1
    args.append("".join(current).strip())
    return args


def strip_outer_parens(text: str) -> str:
    text = text.strip()
    while text.startswith("(") and text.endswith(")"):
        depth = 0
        ok = True
        for index, char in enumerate(text):
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
                if depth == 0 and index != len(text) - 1:
                    ok = False
                    break
        if not ok:
            break
        text = text[1:-1].strip()
    return text


def normalize_sheet_name(name: str) -> str:
    name = name.strip()
    if name.startswith("'") and name.endswith("'"):
        name = name[1:-1].replace("''", "'")
    return name


def cell_to_indexes(addr: str) -> tuple[int, int]:
    match = re.fullmatch(r"\$?([A-Z]{1,3})\$?(\d+)", addr)
    if not match:
        raise ValueError(f"Unsupported cell address: {addr}")
    col, row = match.groups()
    return int(row), column_index_from_string(col)


@dataclass(frozen=True)
class CellRef:
    sheet: str
    row: int
    col: int


class FormulaEvaluator:
    ref_pattern = re.compile(
        r"(?:(?P<sheet>'[^']+'|[^!]+)!)?"
        r"(?P<cell>\$?[A-Z]{1,3}\$?\d+)"
    )
    range_pattern = re.compile(
        r"(?:(?P<sheet>'[^']+'|[^!]+)!)?"
        r"(?P<start>\$?[A-Z]{1,3}\$?\d+):(?P<end>\$?[A-Z]{1,3}\$?\d+)"
    )

    def __init__(self, workbook_path: Path):
        self.formula_wb = load_workbook(workbook_path, data_only=False)
        self.cached_wb = load_workbook(workbook_path, data_only=True)
        self.cache: dict[CellRef, Any] = {}

    def cell_value(self, sheet_name: str, row: int, col: int) -> Any:
        ref = CellRef(sheet_name, row, col)
        if ref in self.cache:
            return self.cache[ref]

        ws = self.formula_wb[sheet_name]
        cell = ws.cell(row, col)
        if isinstance(cell, MergedCell):
            return ""

        raw = cell.value
        if not is_formula(raw):
            self.cache[ref] = raw
            return raw

        cached = self.cached_wb[sheet_name].cell(row, col).value
        if cached not in (None, ""):
            self.cache[ref] = cached
            return cached

        try:
            value = self.eval_expr(str(raw)[1:], sheet_name)
        except Exception:
            value = raw
        self.cache[ref] = value
        return value

    def eval_expr(self, expr: str, sheet_name: str) -> Any:
        expr = strip_outer_parens(expr.strip())
        upper = expr.upper()

        if expr == '""':
            return ""
        if re.fullmatch(r"-?\d+(\.\d+)?", expr):
            number = float(expr) if "." in expr else int(expr)
            return number

        additive = self.split_top_level(expr, ["+", "-"])
        if additive and len(additive) > 1:
            total = self.numeric(self.eval_expr(additive[0][1], sheet_name))
            for op, part in additive[1:]:
                value = self.numeric(self.eval_expr(part, sheet_name))
                total = total + value if op == "+" else total - value
            return total

        if upper.startswith("IF(") and expr.endswith(")"):
            args = split_args(expr[3:-1])
            if len(args) != 3:
                raise ValueError("Unsupported IF formula")
            return self.eval_expr(args[1], sheet_name) if self.truthy(args[0], sheet_name) else self.eval_expr(args[2], sheet_name)

        if upper.startswith("AND(") and expr.endswith(")"):
            return all(self.truthy(arg, sheet_name) for arg in split_args(expr[4:-1]))

        if upper.startswith("SUM(") and expr.endswith(")"):
            total = 0
            for arg in split_args(expr[4:-1]):
                if self.range_pattern.fullmatch(arg.strip()):
                    total += sum(self.numeric(value) for value in self.range_values(arg, sheet_name))
                else:
                    total += self.numeric(self.eval_expr(arg, sheet_name))
            return total

        if upper.startswith("SUMPRODUCT(") and expr.endswith(")"):
            return self.eval_sumproduct(expr[11:-1], sheet_name)

        comparison = self.find_comparison(expr)
        if comparison:
            left, op, right = comparison
            left_value = self.eval_expr(left, sheet_name)
            right_value = self.eval_expr(right, sheet_name)
            if left_value is None and right_value == "":
                left_value = ""
            if right_value is None and left_value == "":
                right_value = ""
            if op == "<>":
                return left_value != right_value
            if op == ">=":
                return self.numeric(left_value) >= self.numeric(right_value)
            if op == "<=":
                return self.numeric(left_value) <= self.numeric(right_value)
            if op == ">":
                return self.numeric(left_value) > self.numeric(right_value)
            if op == "<":
                return self.numeric(left_value) < self.numeric(right_value)
            if op == "=":
                return left_value == right_value

        ref_match = self.ref_pattern.fullmatch(expr)
        if ref_match:
            target_sheet = normalize_sheet_name(ref_match.group("sheet") or sheet_name)
            row, col = cell_to_indexes(ref_match.group("cell"))
            return self.cell_value(target_sheet, row, col)

        return self.eval_arithmetic(expr, sheet_name)

    def eval_sumproduct(self, inner: str, sheet_name: str) -> int:
        inner = strip_outer_parens(inner)
        match = re.fullmatch(r"\((.+)>\s*(-?\d+(?:\.\d+)?)\)\s*\*\s*1", inner)
        if not match:
            raise ValueError("Unsupported SUMPRODUCT formula")
        range_ref, threshold = match.groups()
        limit = float(threshold)
        return sum(1 for value in self.range_values(range_ref.strip(), sheet_name) if self.numeric(value) > limit)

    def split_top_level(self, expr: str, operators: list[str]) -> list[tuple[str, str]] | None:
        parts: list[tuple[str, str]] = []
        current: list[str] = []
        depth = 0
        in_string = False
        current_op = "+"
        for index, char in enumerate(expr):
            if char == '"':
                in_string = not in_string
                current.append(char)
                continue
            if in_string:
                current.append(char)
                continue
            if char == "(":
                depth += 1
                current.append(char)
                continue
            if char == ")":
                depth -= 1
                current.append(char)
                continue
            if depth == 0 and char in operators and index != 0:
                parts.append((current_op, "".join(current).strip()))
                current = []
                current_op = char
                continue
            current.append(char)
        if parts:
            parts.append((current_op, "".join(current).strip()))
            return parts
        return None

    def range_values(self, range_ref: str, sheet_name: str) -> Iterable[Any]:
        match = self.range_pattern.fullmatch(range_ref.strip())
        if not match:
            raise ValueError(f"Unsupported range: {range_ref}")
        target_sheet = normalize_sheet_name(match.group("sheet") or sheet_name)
        start_row, start_col = cell_to_indexes(match.group("start"))
        end_row, end_col = cell_to_indexes(match.group("end"))
        for row in range(min(start_row, end_row), max(start_row, end_row) + 1):
            for col in range(min(start_col, end_col), max(start_col, end_col) + 1):
                yield self.cell_value(target_sheet, row, col)

    def find_comparison(self, expr: str) -> tuple[str, str, str] | None:
        depth = 0
        in_string = False
        operators = ["<>", ">=", "<=", ">", "<", "="]
        for index, char in enumerate(expr):
            if char == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if char == "(":
                depth += 1
                continue
            if char == ")":
                depth -= 1
                continue
            if depth == 0:
                for op in operators:
                    if expr.startswith(op, index):
                        return expr[:index].strip(), op, expr[index + len(op):].strip()
        return None

    def eval_arithmetic(self, expr: str, sheet_name: str) -> Any:
        def replace_ref(match: re.Match[str]) -> str:
            target_sheet = normalize_sheet_name(match.group("sheet") or sheet_name)
            row, col = cell_to_indexes(match.group("cell"))
            return repr(self.numeric(self.cell_value(target_sheet, row, col)))

        python_expr = self.ref_pattern.sub(replace_ref, expr)
        tree = ast.parse(python_expr, mode="eval")
        return self.safe_eval_ast(tree.body)

    def safe_eval_ast(self, node: ast.AST) -> float:
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return node.value
        if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
            return -self.safe_eval_ast(node.operand)
        if isinstance(node, ast.BinOp):
            left = self.safe_eval_ast(node.left)
            right = self.safe_eval_ast(node.right)
            if isinstance(node.op, ast.Add):
                return left + right
            if isinstance(node.op, ast.Sub):
                return left - right
            if isinstance(node.op, ast.Mult):
                return left * right
            if isinstance(node.op, ast.Div):
                return left / right if right else 0
        raise ValueError("Unsupported arithmetic expression")

    def truthy(self, expr: str, sheet_name: str) -> bool:
        value = self.eval_expr(expr, sheet_name)
        if isinstance(value, str):
            return value != ""
        return bool(value)

    @staticmethod
    def numeric(value: Any) -> float:
        if value in (None, ""):
            return 0
        if isinstance(value, bool):
            return 1 if value else 0
        if isinstance(value, (int, float)):
            return value
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0


def used_bounds(ws, evaluator: FormulaEvaluator) -> tuple[int, int]:
    max_row = 0
    max_col = 0
    for row in range(1, ws.max_row + 1):
        if ws.row_dimensions[row].hidden:
            continue
        for col in range(1, ws.max_column + 1):
            if ws.column_dimensions[get_column_letter(col)].hidden:
                continue
            raw = ws.cell(row, col).value
            if raw is None:
                continue
            value = evaluator.cell_value(ws.title, row, col)
            if value not in (None, ""):
                max_row = max(max_row, row)
                max_col = max(max_col, col)
    return max_row, max_col


def export_workbook(workbook_path: Path, out_dir: Path, include_hidden: bool = False) -> list[dict[str, Any]]:
    evaluator = FormulaEvaluator(workbook_path)
    out_dir.mkdir(parents=True, exist_ok=True)

    manifest: list[dict[str, Any]] = []
    visible_index = 1
    for ws in evaluator.formula_wb.worksheets:
        if ws.sheet_state != "visible" and not include_hidden:
            continue

        max_row, max_col = used_bounds(ws, evaluator)
        csv_name = f"{visible_index:02d}_{safe_filename(ws.title)}.csv"
        csv_path = out_dir / csv_name

        with csv_path.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            for row in range(1, max_row + 1):
                if ws.row_dimensions[row].hidden:
                    continue
                values: list[Any] = []
                for col in range(1, max_col + 1):
                    if ws.column_dimensions[get_column_letter(col)].hidden:
                        continue
                    value = evaluator.cell_value(ws.title, row, col)
                    values.append("" if value is None else value)
                writer.writerow(values)

        manifest.append({
            "sheet": ws.title,
            "sheet_state": ws.sheet_state,
            "csv_file": csv_name,
            "rows": max_row,
            "columns": max_col,
        })
        visible_index += 1

    manifest_path = out_dir / "manifest.csv"
    with manifest_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=["sheet", "sheet_state", "csv_file", "rows", "columns"])
        writer.writeheader()
        writer.writerows(manifest)

    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Export each worksheet from an xlsx workbook to separate CSV files.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument("--include-hidden", action="store_true", help="Also export hidden worksheets.")
    args = parser.parse_args()

    manifest = export_workbook(args.workbook.resolve(), args.out_dir.resolve(), args.include_hidden)
    print(f"Exported {len(manifest)} sheets to {args.out_dir.resolve()}")
    for item in manifest:
        print(f"- {item['sheet']} -> {item['csv_file']} ({item['rows']}x{item['columns']})")


if __name__ == "__main__":
    main()
