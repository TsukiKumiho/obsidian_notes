"""
全量习题册分析——复刻 习题册数据分析.ipynb 的分析模式（无 pandas 依赖）

生成图表类型：
  1. 各科/各题册 小节正确率热力图
  2. 章节与分布分析 (4-panel: 正确率bar, 题量bar, 分布hist, 散点)
  3. 横向对比 (全部题册正确率bar)
  4. 计网真题vs自编
  5. 计网进度时间线
"""
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict
import os, re

plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

XLSX = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
FIG_DIR = r'C:\Users\34406\Documents\Obsidian Vault\数据分析\figures'
os.makedirs(FIG_DIR, exist_ok=True)

WB = load_workbook(XLSX)

def read_408_sheet(idx):
    """Parse a 408 sheet: list of (section_name, chapter_name, ch_no, sec_no, 
       total, correct, zt_total, zt_correct, date)"""
    ws = WB.worksheets[idx]
    rows = []
    current_ch = ''
    ch_no = 0
    for r in range(5, ws.max_row+1):
        a = str(ws.cell(row=r, column=1).value or '')
        b = str(ws.cell(row=r, column=2).value or '')
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        f = ws.cell(row=r, column=6).value or 0
        g = ws.cell(row=r, column=7).value or 0
        if '章' in a and '第' in a:
            current_ch = a
            m = re.search(r'第(\d+)章', a)
            if m: ch_no = int(m.group(1))
            continue
        if '合计' in a:
            break
        if isinstance(c, (int,float)) and float(c) > 0:
            sec = re.search(r'\d+\.\d+', a)
            sec_no = sec.group() if sec else ''
            rows.append((a.strip(), current_ch, ch_no, sec_no,
                        float(c), float(d or 0), float(f), float(g), str(b)))
    return rows

def read_880_sheet(idx):
    """Parse 880 sheet: list of (chapter, date, totals, corrects) 
    for 选择+填空 combined."""
    ws = WB.worksheets[idx]
    rows = []
    for r in range(5, ws.max_row+1):
        a = str(ws.cell(row=r, column=1).value or '')
        b = str(ws.cell(row=r, column=2).value or '')
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        if '合计' in a or '总' in a:
            break
        if isinstance(c, (int,float)) and float(c) > 0:
            rows.append((a.strip(), str(b), float(c), float(d or 0)))
    return rows

def read_yxueti_sheet(idx):
    """Parse 严选题 sheet."""
    ws = WB.worksheets[idx]
    rows = []
    for r in range(5, ws.max_row+1):
        a = str(ws.cell(row=r, column=1).value or '')
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        if '合计' in a:
            break
        if isinstance(c, (int,float)) and float(c) > 0:
            rows.append((a.strip(), float(c), float(d or 0)))
    return rows

# =========================
# 1. 热力图
# =========================
def save_heatmap(data, title, filename, n_cols=8):
    """data: list of (section, chapter, total, correct, rate)"""
    if not data: return
    chapters = sorted(set(d[1] for d in data), key=lambda x: int(re.search(r'\d+',x).group()) if re.search(r'\d+',x) else 0)
    ch_sections = defaultdict(list)
    for d in data:
        ch_sections[d[1]].append(d)
    
    max_secs = max(len(v) for v in ch_sections.values())
    n_ch = len(chapters)
    
    fig_w = max(12, max_secs * 1.2)
    fig_h = max(5, n_ch * 0.8)
    
    grid = np.full((n_ch, max_secs), np.nan)
    labels = np.full((n_ch, max_secs), '', dtype=object)
    
    for i, ch in enumerate(chapters):
        for j, sec in enumerate(ch_sections[ch]):
            if sec[2] > 0:
                grid[i,j] = sec[4]  # rate = correct/total
                labels[i,j] = f'{sec[0][:8]}\n{sec[4]*100:.0f}%'
    
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    im = ax.imshow(grid, cmap='RdYlGn', vmin=0.4, vmax=1.0, aspect='auto')
    for i in range(n_ch):
        for j in range(max_secs):
            if labels[i,j]:
                ax.text(j, i, labels[i,j], ha='center', va='center', fontsize=7 if max_secs>10 else 9)
    
    ax.set_yticks(range(n_ch))
    ax.set_yticklabels([ch[:15] for ch in chapters], fontsize=9)
    ax.set_xticks([])
    ax.set_title(title, fontsize=14, fontweight='bold')
    plt.colorbar(im, ax=ax, label='正确率')
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, filename), dpi=150, bbox_inches='tight')
    plt.close()
    print(f'  Heatmap: {filename}')

# =========================
# 2. 四面板章节分析
# =========================
def save_chapter_analysis(data, title, filename):
    """4-panel: rate bar, volume bar, rate distribution, scatter"""
    if not data: return
    chapters = sorted(set(d[1] for d in data), key=lambda x: int(re.search(r'\d+',x).group()) if re.search(r'\d+',x) else 0)
    ch_data = {}
    for ch in chapters:
        secs = [d for d in data if d[1]==ch]
        total = sum(s[2] for s in secs)
        correct = sum(s[3] for s in secs)
        ch_data[ch] = {'total':total, 'correct':correct, 'rate':correct/total if total>0 else 0, 'sections':secs}
    
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    ch_labels = [ch[:12] for ch in chapters]
    rates = [ch_data[ch]['rate']*100 for ch in chapters]
    totals = [ch_data[ch]['total'] for ch in chapters]
    corrects = [ch_data[ch]['correct'] for ch in chapters]
    wrongs = [t-c for t,c in zip(totals, corrects)]
    
    # Panel 1: Chapter accuracy
    colors1 = ['#2F5496' if r>80 else '#C55A11' if r>60 else '#C00000' for r in rates[:-2]] + ['#2F5496','#C55A11']
    colors1 = ['#2F5496' if r>80 else '#C55A11' if r>60 else '#C00000' for r in rates]
    axes[0,0].barh(range(len(ch_labels)), rates, color=colors1, edgecolor='white')
    axes[0,0].set_yticks(range(len(ch_labels)))
    axes[0,0].set_yticklabels(ch_labels, fontsize=9)
    for i, r in enumerate(rates):
        axes[0,0].text(r+1, i, f'{r:.1f}%', va='center', fontsize=8)
    axes[0,0].set_title('各章节正确率', fontsize=13, fontweight='bold')
    axes[0,0].set_xlim(0, 105)
    
    # Panel 2: Question volume (stacked: correct + wrong)
    axes[0,1].barh(range(len(ch_labels)), corrects, color='#2F5496', label='正确', edgecolor='white')
    axes[0,1].barh(range(len(ch_labels)), wrongs, left=corrects, color='#C00000', label='错误', edgecolor='white')
    axes[0,1].set_yticks(range(len(ch_labels)))
    axes[0,1].set_yticklabels(ch_labels, fontsize=9)
    axes[0,1].set_title('各章节题量（正确/错误）', fontsize=13, fontweight='bold')
    axes[0,1].legend(fontsize=9)
    
    # Panel 3: Section rate distribution histogram
    sec_rates = [d[4]*100 for d in data]
    axes[1,0].hist(sec_rates, bins=15, color='#2F5496', edgecolor='white', alpha=0.8)
    axes[1,0].axvline(x=np.mean(sec_rates), color='red', linestyle='--', label=f'平均 {np.mean(sec_rates):.1f}%')
    axes[1,0].set_title('小节正确率分布', fontsize=13, fontweight='bold')
    axes[1,0].set_xlabel('正确率 (%)')
    axes[1,0].legend()
    
    # Panel 4: Volume vs accuracy scatter
    sec_totals = [d[2] for d in data]
    axes[1,1].scatter(sec_totals, sec_rates, alpha=0.6, s=60, c='#2F5496')
    axes[1,1].set_title('小节题量 vs 正确率', fontsize=13, fontweight='bold')
    axes[1,1].set_xlabel('题量')
    axes[1,1].set_ylabel('正确率 (%)')
    
    plt.suptitle(title, fontsize=15, fontweight='bold', y=1.01)
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, filename), dpi=150, bbox_inches='tight')
    plt.close()
    print(f'  Chapter analysis: {filename}')

# =========================
# 3. 横向对比
# =========================
def save_comparison(data_dict, filename):
    """data_dict: {name: (total, correct)}"""
    fig, ax = plt.subplots(figsize=(12, 6))
    names = list(data_dict.keys())
    rates = [c/t*100 if t>0 else 0 for t,c in [data_dict[n] for n in names]]
    totals = [data_dict[n][0] for n in names]
    
    x = np.arange(len(names))
    colors = ['#2F5496','#C55A11','#548235','#7030A0','#BF8F00','#C00000','#31859C'][:len(names)]
    bars = ax.bar(x, rates, color=colors, edgecolor='white', linewidth=2)
    for bar, rate, total in zip(bars, rates, totals):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
                f'{rate:.1f}%\n({total}题)', ha='center', fontsize=9, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(names, fontsize=11)
    ax.set_ylabel('正确率 (%)', fontsize=12)
    ax.set_title('各习题册正确率横向对比', fontsize=15, fontweight='bold')
    ax.set_ylim(0, 100)
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.yaxis.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, filename), dpi=150, bbox_inches='tight')
    plt.close()
    print(f'  Comparison: {filename}')

# =========================
# MAIN
# =========================
print('Reading data...')

# 408 sheets
all_408 = {}
for idx, name in [(3,'DS'),(4,'CO'),(5,'OS'),(6,'CN')]:
    rows = read_408_sheet(idx)
    all_408[name] = rows
    # Convert to section data format for heatmap/chapter
    # (section_label, chapter_name, total, correct, rate)
    sec_data = [(r[3] or r[0][:8], r[1], r[4], r[5], r[5]/r[4] if r[4]>0 else 0) for r in rows]
    save_heatmap(sec_data, f'{name} 小节正确率热力图', f'heatmap_{name}.png')
    

# Chapter analysis for each 408 subject
for name in ['DS','CO','OS','CN']:
    ch_analysis_data = [(r[0][:8], r[1], r[4], r[5], r[5]/r[4] if r[4]>0 else 0) for r in all_408[name]]
    save_chapter_analysis(ch_analysis_data, f'{name} 章节分析', f'chapter_{name}.png')

# 880
print('Reading 880...')
rows_880 = read_880_sheet(2)
sec_880 = [(r[0], r[0], r[2], r[3], r[3]/r[2] if r[2]>0 else 0) for r in rows_880 if '章' in r[0]]
save_heatmap(sec_880, '李林880 章节正确率', 'heatmap_880.png')
save_chapter_analysis([(r[0], r[0], r[2], r[3], r[3]/r[2] if r[2]>0 else 0) for r in rows_880 if '章' in r[0]],
                       '李林880 章节分析', 'chapter_880.png')

# 严选题
print('Reading 严选题...')
rows_yx = read_yxueti_sheet(1)
sec_yx = [(r[0], r[0], r[1], r[2], r[2]/r[1] if r[1]>0 else 0) for r in rows_yx if '章' in r[0]]
save_heatmap(sec_yx, '严选题 章节正确率', 'heatmap_yxueti.png')
save_chapter_analysis([(r[0], r[0], r[1], r[2], r[2]/r[1] if r[1]>0 else 0) for r in rows_yx if '章' in r[0]],
                       '严选题 章节分析', 'chapter_yxueti.png')

# Cross-comparison
print('Cross comparison...')
comp = {}
for name in ['DS','CO','OS','CN']:
    t = sum(r[4] for r in all_408[name])
    c = sum(r[5] for r in all_408[name])
    comp[f'408_{name}'] = (t, c)
# 880
t = sum(r[2] for r in rows_880)
c = sum(r[3] for r in rows_880)
comp['880'] = (t, c)
# 严选题
t = sum(r[1] for r in rows_yx)
c = sum(r[2] for r in rows_yx)
comp['严选题'] = (t, c)
save_comparison(comp, 'all_comparison.png')

# CN 真题 vs 自编
print('CN 真题vs自编...')
cn_zt = []
for r in all_408['CN']:
    if r[6] > 0:  # 真题有题量
        cn_zt.append((r[0][:15], r[6], r[7], r[7]/r[6]*100 if r[6]>0 else 0))

if cn_zt:
    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(len(cn_zt)); w = 0.35
    zt_rates = [d[3] for d in cn_zt]
    zb_rates = []
    for d, zt in zip(all_408['CN'], cn_zt):
        zt_t, zt_c = zt[1], zt[2]
        zb_t, zb_c = d[4]-zt_t, d[5]-zt_c
        zb_rates.append(zb_c/zb_t*100 if zb_t>0 else 0)
    
    ax.bar(x-w/2, zt_rates, w, label='真题', color='#2F5496', edgecolor='white')
    ax.bar(x+w/2, zb_rates, w, label='自编', color='#C55A11', edgecolor='white')
    ax.set_xticks(x)
    ax.set_xticklabels([d[0] for d in cn_zt], fontsize=8, rotation=30, ha='right')
    ax.set_title('计网 真题 vs 自编 正确率对比', fontsize=14, fontweight='bold')
    ax.legend()
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.yaxis.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, 'cn_zt_zb.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f'  CN zt-zb: cn_zt_zb.png')

# CN progress
print('CN progress...')
cn_dates = [(r[8], r[4]) for r in all_408['CN'] if r[8]]
date_sum = defaultdict(float)
for d, t in cn_dates:
    if d and len(d)==3:
        date_sum[d] += t

if date_sum:
    sorted_dates = sorted(date_sum.keys())
    cum = 0; xs, ys, xlabs = [], [], []
    for d in sorted_dates:
        cum += date_sum[d]
        xs.append(len(xs))
        ys.append(cum)
        xlabs.append(f'6/{int(d[1:])}')
    
    fig, ax = plt.subplots(figsize=(9, 4))
    ax.fill_between(xs, 0, ys, alpha=0.2, color='#2F5496')
    ax.plot(xs, ys, 'o-', color='#2F5496', linewidth=2.5, markersize=8)
    for i, y in enumerate(ys):
        ax.text(i, y+8, str(int(y)), ha='center', fontsize=9, fontweight='bold')
    ax.set_xticks(xs); ax.set_xticklabels(xlabs)
    ax.set_title('计网 累计做题进度', fontsize=14, fontweight='bold')
    ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
    ax.yaxis.grid(True, alpha=0.3)
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, 'cn_progress.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print(f'  CN progress: cn_progress.png')

print('\nAll analysis charts generated to 数据分析/figures/')
