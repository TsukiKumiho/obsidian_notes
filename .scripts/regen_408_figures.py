"""Regenerate pipeline-standard 408 figures — FIXED"""
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from collections import defaultdict
import os, re

plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus'] = False

XLSX = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
FIG = r'C:\Users\34406\Documents\Obsidian Vault\数据分析\figures'
os.makedirs(FIG, exist_ok=True)

WB = load_workbook(XLSX)
SUBJ = {'DS':'数据结构','CO':'计组','OS':'操作系统','CN':'计网'}
COLORS = {'DS':'#2F5496','CO':'#C55A11','OS':'#548235','CN':'#7030A0'}

def read_408(idx):
    ws = WB.worksheets[idx]
    rows = []
    ch = ''; ch_no = 0
    for r in range(5, ws.max_row+1):
        a = str(ws.cell(row=r, column=1).value or '')
        b = str(ws.cell(row=r, column=2).value or '')
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        f = ws.cell(row=r, column=6).value or 0
        g = ws.cell(row=r, column=7).value or 0
        if '章' in a and '第' in a:
            ch = a; m = re.search(r'第(\d+)章', a)
            if m: ch_no = int(m.group(1))
            continue
        if '合计' in a: break
        if isinstance(c,(int,float)) and float(c)>0:
            rows.append((a.strip(), ch, ch_no, float(c), float(d or 0), float(f), float(g)))
    return rows

all_data = {}
for idx, name in [(3,'DS'),(4,'CO'),(5,'OS'),(6,'CN')]:
    all_data[name] = read_408(idx)

# Build flat section list
all_secs = []  # (section_name, chapter_name, ch_no, subject, total, correct, zt_t, zt_c)
for name in ['DS','CO','OS','CN']:
    for r in all_data[name]:
        total = r[3]; correct = r[4]
        if total > 0:
            all_secs.append((r[0], r[1], r[2], name, total, correct, r[5], r[6]))

# ===== 408_overview.png =====
fig, ax = plt.subplots(figsize=(9, 5.5))
names = ['DS','CO','OS','CN']; full = ['数据结构','计算机组成原理','操作系统','计算机网络']
rates = []
for n in names:
    t = sum(r[3] for r in all_data[n])
    c = sum(r[4] for r in all_data[n])
    rates.append(c/t*100 if t>0 else 0)
bars = ax.bar(full, rates, color=[COLORS[n] for n in names], edgecolor='white', linewidth=2)
for bar, r in zip(bars, rates):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+1,
            f'{r:.1f}%', ha='center', fontsize=14, fontweight='bold')
ax.set_ylim(0, 100); ax.set_ylabel('正确率 (%)', fontsize=13)
ax.set_title('408 四科一刷正确率总览', fontsize=16, fontweight='bold')
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_overview.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_overview.png')

# ===== 408_heatmap.png =====
fig, ax = plt.subplots(figsize=(20, 4.5))
sec_names = [f'{s[3]}:{s[0][:8]}' for s in all_secs]
sec_rates = [s[5]/s[4]*100 for s in all_secs]
x = np.arange(len(sec_rates))
colors2 = ['#2F5496' if r>80 else '#C55A11' if r>60 else '#C00000' for r in sec_rates]
ax.bar(x, sec_rates, color=colors2, edgecolor='white', linewidth=0.5, width=0.85)
last_subj = ''
for i, s in enumerate(all_secs):
    if s[3] != last_subj:
        ax.axvline(x=i-0.5, color='black', linewidth=1.5, alpha=0.5)
        ax.text(i-0.25, 102, SUBJ[s[3]], fontsize=11, fontweight='bold', 
                color=COLORS[s[3]], ha='center', va='bottom')
        last_subj = s[3]
ax.set_xticks([])
ax.set_ylabel('正确率 (%)', fontsize=12)
ax.set_title('408 四科 小节正确率总览', fontsize=15, fontweight='bold')
ax.set_ylim(0, 108)
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_heatmap.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_heatmap.png')

# ===== 408_chapter_ranking.png (FIXED: proper rate = sum(correct)/sum(total)) =====
ch_agg = {}  # key: (subject, chapter_name) -> [total, correct]
for s in all_secs:
    key = (s[3], s[1])
    if key not in ch_agg:
        ch_agg[key] = [0,0]
    ch_agg[key][0] += s[4]
    ch_agg[key][1] += s[5]

ch_list = [(k[0], k[1], v[0], v[1], v[1]/v[0]*100) for k,v in ch_agg.items()]
ch_list.sort(key=lambda x: -x[4])

fig, ax = plt.subplots(figsize=(10, 10))
clabels = [f'{SUBJ[c[0]]} {c[1][:15]}' for c in ch_list]
crates = [c[4] for c in ch_list]
ccolors = [COLORS[c[0]] for c in ch_list]
ax.barh(range(len(clabels)), crates, color=ccolors, edgecolor='white', alpha=0.85)
ax.set_yticks(range(len(clabels))); ax.set_yticklabels(clabels, fontsize=9)
for i, c in enumerate(ch_list):
    ax.text(crates[i]+0.5, i, f'{crates[i]:.1f}% ({int(c[3])}/{int(c[2])})', va='center', fontsize=8)
ax.set_xlim(0, 105); ax.set_xlabel('正确率 (%)')
ax.set_title('各章节正确率排名', fontsize=15, fontweight='bold')
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_chapter_ranking.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_chapter_ranking.png')

# ===== 408_weak_sections.png (FIXED: label = "OS 文件系统" style) =====
sec_list = [(s[3], s[1], s[0], s[4], s[5], s[5]/s[4]*100) for s in all_secs]
sec_list.sort(key=lambda x: x[5])
weak = sec_list[:15]
fig, ax = plt.subplots(figsize=(11, 5.5))
wlabels = [f'{SUBJ[w[0]]} {w[1][:3]} {w[2][:15]}' for w in weak]
wrates = [w[5] for w in weak]
wcolors = ['#C00000' if r<50 else '#C55A11' if r<70 else '#BF8F00' for r in wrates]
ax.barh(range(len(wlabels)), wrates, color=wcolors, edgecolor='white')
ax.set_yticks(range(len(wlabels))); ax.set_yticklabels(wlabels, fontsize=9)
for i, w in enumerate(weak):
    ax.text(w[5]+0.5, i, f'{w[5]:.1f}% ({int(w[4])}/{int(w[3])})', va='center', fontsize=7.5)
ax.set_xlim(0, 85); ax.set_xlabel('正确率 (%)')
ax.set_title('薄弱小节 Top 15', fontsize=15, fontweight='bold')
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_weak_sections.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_weak_sections.png')

# ===== 408_zhenti_vs_zibian.png =====
zt_rates, zb_rates = [], []
for name in ['DS','CO','OS','CN']:
    zt_t = sum(r[5] for r in all_data[name])
    zt_c = sum(r[6] for r in all_data[name])
    zb_t = sum(r[3]-r[5] for r in all_data[name])
    zb_c = sum(r[4]-r[6] for r in all_data[name])
    zt_rates.append(zt_c/zt_t*100 if zt_t>0 else 0)
    zb_rates.append(zb_c/zb_t*100 if zb_t>0 else 0)

fig, ax = plt.subplots(figsize=(9, 5.5))
x = np.arange(4); w = 0.35
ax.bar(x-w/2, zt_rates, w, label='真题', color='#2F5496', edgecolor='white')
ax.bar(x+w/2, zb_rates, w, label='自编', color='#C55A11', edgecolor='white')
for i in range(4):
    ax.text(i-w/2, zt_rates[i]+1, f'{zt_rates[i]:.1f}%', ha='center', fontsize=10, fontweight='bold')
    ax.text(i+w/2, zb_rates[i]+1, f'{zb_rates[i]:.1f}%', ha='center', fontsize=10, fontweight='bold')
ax.set_xticks(x); ax.set_xticklabels(full, fontsize=13)
ax.set_ylabel('正确率 (%)'); ax.set_ylim(0, 100)
ax.set_title('408 四科 真题 vs 自编', fontsize=15, fontweight='bold')
ax.legend(fontsize=12)
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_zhenti_vs_zibian.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_zhenti_vs_zibian.png')

# ===== 408_wrong_count.png (FIXED: label = "OS 文件系统" style) =====
wrong_list = [(s[3], s[1], s[0], s[4]-s[5], s[4], (1-s[5]/s[4])*100) for s in all_secs if s[4]>0]
wrong_list.sort(key=lambda x: -x[3])
top_wrong = wrong_list[:15]

fig, ax = plt.subplots(figsize=(11, 5.5))
wl2 = [f'{SUBJ[w[0]]} {w[1][:3]} {w[2][:15]}' for w in top_wrong]
wv = [w[3] for w in top_wrong]
wcolors2 = ['#C00000' if w>15 else '#C55A11' if w>8 else '#BF8F00' for w in wv]
ax.barh(range(len(wl2)), wv, color=wcolors2, edgecolor='white', alpha=0.85)
ax.set_yticks(range(len(wl2))); ax.set_yticklabels(wl2, fontsize=9)
for i, w in enumerate(top_wrong):
    ax.text(w[3]+0.3, i, f'{int(w[3])}题 /{int(w[4])} ({w[5]:.0f}%)', va='center', fontsize=7.5)
ax.set_xlabel('错题数'); ax.set_title('错题最多小节 Top 15', fontsize=15, fontweight='bold')
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_wrong_count.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_wrong_count.png')

# ===== 408_distribution.png (FIXED: split by subject, 4 panels) =====
fig, axes = plt.subplots(2, 2, figsize=(14, 11))
for idx, name in enumerate(['DS','CO','OS','CN']):
    ax = axes[idx//2][idx%2]
    sub_secs = [s for s in all_secs if s[3]==name]
    sub_rates = [s[5]/s[4]*100 for s in sub_secs]
    sub_totals = [s[4] for s in sub_secs]
    
    ax.hist(sub_rates, bins=12, color=COLORS[name], edgecolor='white', alpha=0.8)
    avg = np.mean(sub_rates) if sub_rates else 0
    ax.axvline(avg, color='red', ls='--', lw=2, label=f'平均 {avg:.1f}%')
    ax.set_title(f'{SUBJ[name]} 正确率分布', fontsize=13, fontweight='bold')
    ax.set_xlabel('正确率 (%)'); ax.set_xlim(0, 105)
    ax.legend(fontsize=9)
    # Add scatter as twin
    ax2 = ax.twiny()
    ax2.scatter(sub_totals, sub_rates, alpha=0.3, s=40, c=COLORS[name])
    ax2.set_xlim(0, max(sub_totals)*1.1 if sub_totals else 50)
    ax2.set_xlabel('  题量 →', fontsize=8, color='gray')

plt.suptitle('各科目小节正确率分布 & 题量散点', fontsize=15, fontweight='bold', y=1.01)
plt.tight_layout()
fig.savefig(os.path.join(FIG, '408_distribution.png'), dpi=150, bbox_inches='tight')
plt.close()
print('[OK] 408_distribution.png')

print('\nAll done.')
