"""
重建李林880选填区：参考严选题列布局 → 基础篇(选/填) + 综合篇(选/填) + 合计
解答区完全不动，逐行复制原数据。不使用insert_cols。
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
wb = openpyxl.load_workbook(SRC)

# Read ALL old cells
ws_old = wb.worksheets[2]
old_title = ws_old.title
old_cells = {}
for r in range(1, ws_old.max_row + 1):
    for c in range(1, ws_old.max_column + 1):
        cell = ws_old.cell(row=r, column=c)
        old_cells[(r, c)] = (cell.value, cell.number_format)

wb.remove(ws_old)
ws = wb.create_sheet(title=old_title, index=2)

# Styles
thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
sfill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
hfont = Font(bold=True, size=9, color='FFFFFFFF')
bfont = Font(bold=True, size=10)
dfont = Font(size=10)
bold_white = Font(bold=True, size=9, color='FFFFFFFF')

def w(row, col, value, font=None, fill=None, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or dfont
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    if fmt: c.number_format = fmt
    return c

# ===== 选填区：14列，参考严选题布局 =====
# A:章节 B:日期
# C-E:基础-选择 F-H:基础-填空
# I-K:综合-选择 L-N:综合-填空
# (no 扩展/解答 here — 解答在下面独立区域)

# R1: Title
ws.merge_cells('A1:N1'); w(1, 1, '李林880 （选择&填空）习题正确率记录', Font(bold=True, size=14))

# R3-5: Headers (like 严选题 style — merged section labels → sub-labels → column names)
ws.merge_cells('C3:H3'); w(3, 3, '基础篇', bfont, sfill)
ws.merge_cells('I3:N3'); w(3, 9, '综合篇', bfont, sfill)

for sc, label in [(3,'选择'),(6,'填空'),(9,'选择'),(12,'填空')]:
    ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=sc+2)
    w(4, sc, label, bfont, sfill)

for col, hdr in [(1,'章节'),(2,'日期'),
    (3,'总题数'),(4,'正确数'),(5,'正确率'),(6,'总题数'),(7,'正确数'),(8,'正确率'),
    (9,'总题数'),(10,'正确数'),(11,'正确率'),(12,'总题数'),(13,'正确数'),(14,'正确率')]:
    w(5, col, hdr, hfont, hfill)

def rate_f(tc, cc, r):
    return f'=IF(AND({tc}{r}<>"",{cc}{r}<>""),{cc}{r}/{tc}{r},"")'

gaoshu = [
    ('第一章 函数、极限、连续', '618'), ('第二章 一元函数微分学及其应用', '626'),
    ('第三章 一元函数积分学及其应用', '704'), ('第四章 空间解析几何', ''),
    ('第五章 多元函数微分学及其应用', ''), ('第六章 重积分及其应用', ''),
    ('第七章 微分方程及其应用', ''), ('第八章 无穷级数', ''),
    ('第九章 曲线积分与曲面积分', ''),
]
xiandai = [
    ('第十章 行列式', ''), ('第十一章 矩阵', ''), ('第十二章 向量', ''),
    ('第十三章 线性方程组', ''), ('第十四章 相似矩阵', ''), ('第十五章 二次型', ''),
]
gailv = [
    ('第十六章 随机事件及其概率', ''), ('第十七章 随机变量及其分布', ''),
    ('第十八章 多维随机变量及其分布', ''), ('第十九章 随机变量的数字特征', ''),
    ('第二十章 大数定律与中心极限定理', ''), ('第二十一章 数理统计的基本概念', ''),
    ('第二十二章 参数估计', ''), ('第二十三章 假设检验', ''),
]

def write_subject(ws, row, chs, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
    w(row, 1, label, bfont, sfill)
    row += 1
    first = row
    for ch, date in chs:
        w(row, 1, ch); w(row, 2, date)
        for tc,cc,rc in [(3,4,5),(6,7,8),(9,10,11),(12,13,14)]:
            w(row, rc, rate_f(get_column_letter(tc), get_column_letter(cc), row))
        row += 1
    last = row - 1
    # Sum row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    w(row, 1, f'{label} 合计', bfont, sfill)
    for col in [3,4,6,7,9,10,12,13]:
        cl = get_column_letter(col)
        w(row, col, f'=SUM({cl}{first}:{cl}{last})')
    for tc,cc,rc in [(3,4,5),(6,7,8),(9,10,11),(12,13,14)]:
        w(row, rc, rate_f(get_column_letter(tc), get_column_letter(cc), row))
    # fill sum row columns with section fill
    for c in range(3, 15):
        ws.cell(row=row, column=c).fill = sfill
    return row

row = 6
gs_sum = write_subject(ws, row, gaoshu, '高等数学')
row = gs_sum + 2
xd_sum = write_subject(ws, row, xiandai, '线性代数')
row = xd_sum + 2
gl_sum = write_subject(ws, row, gailv, '概率论与数理统计')
row = gl_sum + 2

# Grand total
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
w(row, 1, '总 计', bold_white, hfill)
for col in [3,4,6,7,9,10,12,13]:
    cl = get_column_letter(col)
    w(row, col, f'={cl}{gs_sum}+{cl}{xd_sum}+{cl}{gl_sum}', bold_white, hfill)
for tc,cc,rc in [(3,4,5),(6,7,8),(9,10,11),(12,13,14)]:
    w(row, rc, rate_f(get_column_letter(tc), get_column_letter(cc), row), bold_white, hfill)
for c in range(3, 15): ws.cell(row=row, column=c).fill = hfill
row += 3

# ===== 解答区：逐行逐列复制原数据（8列），不做任何修改 =====
# Old 解答 starts around row 38
# Copy from old sheet: rows where column A has content, col 1-8
old_rows = sorted(set(r for (r, c) in old_cells if c == 1))
old_解答_rows = [r for r in old_rows if r >= 37]  # 解答 section starts at ~row 38+

for old_r in old_解答_rows:
    for old_c in range(1, 9):
        key = (old_r, old_c)
        if key in old_cells:
            val, nf = old_cells[key]
            if val is not None:
                c = ws.cell(row=row, column=old_c, value=val)
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
                if nf: c.number_format = nf
    row += 1

# Column widths
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 8
for c in range(3, 15):
    ws.column_dimensions[get_column_letter(c)].width = 7

wb.save(SRC)
print('Done. 选填区14列(严选题风格), 解答区逐行复制原样8列.')
