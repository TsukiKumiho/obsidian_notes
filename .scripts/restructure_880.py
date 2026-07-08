import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx')
ws_old = wb.worksheets[2]
old_title = ws_old.title
wb.remove(ws_old)
ws = wb.create_sheet(title=old_title, index=2)

thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hfill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
sfill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
hfont = Font(bold=True, size=9, color='FFFFFFFF')
bfont = Font(bold=True, size=10)
dfont = Font(size=10)

# Row 1: Title
ws.merge_cells('A1:T1')
c = ws.cell(row=1, column=1, value='李林880 （选择&填空）习题正确率记录')
c.font = Font(bold=True, size=14)
c.alignment = Alignment(horizontal='center')

# Row 2: Section headers
# 基础篇 (C-H), 综合篇 (I-N), 合计 (O-T)
ws.merge_cells('C2:H2')
ws.merge_cells('I2:N2')
ws.merge_cells('O2:T2')
for col, label in [(3, '基础篇'), (9, '综合篇'), (15, '合计')]:
    c = ws.cell(row=2, column=col, value=label)
    c.font = bfont; c.fill = sfill; c.alignment = Alignment(horizontal='center')

# Row 2 sub: 选择/填空 under each section
for col_start, sec in [(3,''),(9,''),(15,'')]:
    ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+2)

# Row 3: Column headers
col_specs = {
    'A3': '章节', 'B3': '日期',
    'C3': '总题数', 'D3': '正确数', 'E3': '正确率',
    'F3': '总题数', 'G3': '正确数', 'H3': '正确率',
    'I3': '总题数', 'J3': '正确数', 'K3': '正确率',
    'L3': '总题数', 'M3': '正确数', 'N3': '正确率',
    'O3': '总题数', 'P3': '正确数', 'Q3': '正确率',
    'R3': '总题数', 'S3': '正确数', 'T3': '正确率',
}
for ref, val in col_specs.items():
    c = ws[ref]
    c.value = val
    c.font = hfont; c.fill = hfill
    c.alignment = Alignment(horizontal='center', vertical='center')

# Redo row 2 correctly — put 选择/填空 labels as sub-text
ws.unmerge_cells('C2:H2')
ws.unmerge_cells('I2:N2')
ws.unmerge_cells('O2:T2')
ws.merge_cells('C2:E2')
ws.merge_cells('F2:H2')
ws.merge_cells('I2:K2')
ws.merge_cells('L2:N2')
ws.merge_cells('O2:Q2')
ws.merge_cells('R2:T2')

ws.cell(row=2, column=3, value='选择').font = bfont; ws.cell(row=2, column=3).fill = sfill; ws.cell(row=2, column=3).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=6, value='填空').font = bfont; ws.cell(row=2, column=6).fill = sfill; ws.cell(row=2, column=6).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=9, value='选择').font = bfont; ws.cell(row=2, column=9).fill = sfill; ws.cell(row=2, column=9).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=12, value='填空').font = bfont; ws.cell(row=2, column=12).fill = sfill; ws.cell(row=2, column=12).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=15, value='选择').font = bfont; ws.cell(row=2, column=15).fill = sfill; ws.cell(row=2, column=15).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=18, value='填空').font = bfont; ws.cell(row=2, column=18).fill = sfill; ws.cell(row=2, column=18).alignment = Alignment(horizontal='center')

# Add 基础篇/综合篇/合计 labels above in an extra row
ws.insert_rows(2)
ws.merge_cells('C2:H2')
ws.merge_cells('I2:N2')
ws.merge_cells('O2:T2')
ws.cell(row=2, column=3, value='基础篇').font = bfont; ws.cell(row=2, column=3).fill = sfill; ws.cell(row=2, column=3).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=9, value='综合篇').font = bfont; ws.cell(row=2, column=9).fill = sfill; ws.cell(row=2, column=9).alignment = Alignment(horizontal='center')
ws.cell(row=2, column=15, value='合计').font = bfont; ws.cell(row=2, column=15).fill = sfill; ws.cell(row=2, column=15).alignment = Alignment(horizontal='center')

def rate_f(tc, cc, r):
    return f'=IF(AND({tc}{r}<>\"\",{cc}{r}<>\"\"),{cc}{r}/{tc}{r},\"\")'

gaoshu = [
    ('第一章 函数、极限、连续', '618'),
    ('第二章 一元函数微分学及其应用', '626'),
    ('第三章 一元函数积分学及其应用', '704'),
    ('第四章 空间解析几何', ''),
    ('第五章 多元函数微分学及其应用', ''),
    ('第六章 重积分及其应用', ''),
    ('第七章 微分方程及其应用', ''),
    ('第八章 无穷级数', ''),
    ('第九章 曲线积分与曲面积分', ''),
]
xiandai = [
    ('第十章 行列式', ''), ('第十一章 矩阵', ''), ('第十二章 向量', ''),
    ('第十三章 线性方程组', ''), ('第十四章 相似矩阵', ''), ('第十五章 二次型', ''),
]
gailv = [
    ('第十六章 随机事件及其概率', ''), ('第十七章 随机变量及其分布', ''),
    ('第十八章 多维随机变量及其分布', ''), ('第十九章 随机变量的数字特征', ''),
    ('第二十章 大数定律与中心极限定理', ''), ('第二十一章 数理统计的基本概念', ''),
    ('第二十二章 参数估计', ''), ('第二十三章 假设检验', ''),
]

def write_section(ws, row, chapters, label):
    ws.merge_cells(f'A{row}:T{row}')
    ws.cell(row=row, column=1, value=label).font = bfont
    ws.cell(row=row, column=1).fill = sfill
    row += 1
    first = row
    for ch, date in chapters:
        ws.cell(row=row, column=1, value=ch).font = dfont
        ws.cell(row=row, column=2, value=date).font = dfont
        for rc in ['E','H','K','N','Q','T']:
            tc = get_column_letter(ord(rc)-2)
            cc = get_column_letter(ord(rc)-1)
            ws.cell(row=row, column=ord(rc)-64).value = rate_f(tc, cc, row)
        # 合计-选择 O: C+I, P: D+J
        ws.cell(row=row, column=15).value = f'=C{row}+I{row}'
        ws.cell(row=row, column=16).value = f'=D{row}+J{row}'
        # 合计-填空 R: F+L, S: G+M
        ws.cell(row=row, column=18).value = f'=F{row}+L{row}'
        ws.cell(row=row, column=19).value = f'=G{row}+M{row}'
        for c in range(1, 21):
            ws.cell(row=row, column=c).border = border
        row += 1
    last = row - 1
    # sum row
    ws.merge_cells(f'A{row}:B{row}')
    ws.cell(row=row, column=1, value=f'{label} 合计').font = bfont
    ws.cell(row=row, column=1).fill = sfill
    for col in [3,4,6,7,9,10,12,13,15,16,18,19]:
        cl = get_column_letter(col)
        ws.cell(row=row, column=col).value = f'=SUM({cl}{first}:{cl}{last})'
    for rc in ['E','H','K','N','Q','T']:
        tc = get_column_letter(ord(rc)-2); cc = get_column_letter(ord(rc)-1)
        ws.cell(row=row, column=ord(rc)-64).value = rate_f(tc, cc, row)
    ws.cell(row=row, column=15).value = f'=C{row}+I{row}'
    ws.cell(row=row, column=16).value = f'=D{row}+J{row}'
    ws.cell(row=row, column=18).value = f'=F{row}+L{row}'
    ws.cell(row=row, column=19).value = f'=G{row}+M{row}'
    for c in range(1, 21):
        ws.cell(row=row, column=c).border = border
    return row  # return the sum row

row = 5
gs_sum = write_section(ws, row, gaoshu, '高等数学')
row = gs_sum + 2
xd_sum = write_section(ws, row, xiandai, '线性代数')
row = xd_sum + 2
gl_sum = write_section(ws, row, gailv, '概率论与数理统计')
row = gl_sum + 2

# Grand total
ws.merge_cells(f'A{row}:B{row}')
ws.cell(row=row, column=1, value='总 计').font = Font(bold=True, size=11, color='FFFFFFFF')
ws.cell(row=row, column=1).fill = hfill
for col in [3,4,6,7,9,10,12,13,15,16,18,19]:
    cl = get_column_letter(col)
    ws.cell(row=row, column=col).value = f'={cl}{gs_sum}+{cl}{xd_sum}+{cl}{gl_sum}'
    ws.cell(row=row, column=col).font = Font(bold=True, color='FFFFFFFF')
    ws.cell(row=row, column=col).fill = hfill
for rc in ['E','H','K','N','Q','T']:
    tc = get_column_letter(ord(rc)-2); cc = get_column_letter(ord(rc)-1)
    ws.cell(row=row, column=ord(rc)-64).value = rate_f(tc, cc, row)
    ws.cell(row=row, column=ord(rc)-64).font = Font(bold=True, color='FFFFFFFF')
    ws.cell(row=row, column=ord(rc)-64).fill = hfill
for c in range(1, 21):
    ws.cell(row=row, column=c).border = border

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 8
for c in range(3, 21):
    ws.column_dimensions[get_column_letter(c)].width = 7

wb.save(r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx')
print('Done.')
