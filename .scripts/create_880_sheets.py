"""
生成李林880刷题记录表 + 汇总统计表，并整合到习题册.xlsx
运行依赖: pip install openpyxl
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from copy import copy
import os, re
from collections import defaultdict

# ============================================================
# 配置
# ============================================================
VAULT = r'C:\Users\34406\Documents\Obsidian Vault'
XLSX_PATH = os.path.join(VAULT, '习题册.xlsx')
RECORD_SHEET = '李林880_刷题记录'
SUMMARY_SHEET = '李林880_数据汇总'

CHAPTERS = [
    '第一章 函数、极限、连续',
    '第二章 一元函数微分学及其应用',
    '第三章 一元函数积分学及其应用',
    '第四章 空间解析几何',
    '第五章 多元函数微分学及其应用',
    '第六章 重积分及其应用',
    '第七章 微分方程及其应用',
    '第八章 无穷级数',
    '第九章 曲线积分与曲面积分',
    '第十章 行列式',
    '第十一章 矩阵',
    '第十二章 向量',
    '第十三章 线性方程组',
    '第十四章 相似矩阵',
    '第十五章 二次型',
    '第十六章 随机事件及其概率',
    '第十七章 随机变量及其分布',
    '第十八章 多维随机变量及其分布',
    '第十九章 随机变量的数字特征',
    '第二十章 大数定律与中心极限定理',
    '第二十一章 数理统计的基本概念',
    '第二十二章 参数估计',
    '第二十三章 假设检验',
]

PIAN_MU = ['基础篇', '综合篇', '扩展篇']
TI_XING = ['选择题', '填空题', '解答题']

thin = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
section_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
header_font = Font(bold=True, size=9, color='FFFFFFFF')
bold_font = Font(bold=True, size=10)
data_font = Font(size=10)

# ============================================================
# Sheet 1: 刷题记录
# ============================================================
def create_record_sheet(wb):
    ws = wb.create_sheet(RECORD_SHEET)
    
    # Title
    ws.merge_cells('A1:G1')
    c = ws.cell(row=1, column=1, value='李林880 刷题记录')
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal='center')
    
    # Headers
    headers = ['章节', '篇目', '题型', '日期', '总题目数', '正确题目数', '正确率']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    
    # Column widths
    widths = [28, 10, 10, 12, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    
    # Data validation for 章节
    dv_chapter = DataValidation(type='list', formula1='"' + ','.join(CHAPTERS) + '"', allow_blank=True)
    dv_chapter.error = '请选择有效的章节'
    dv_chapter.errorTitle = '无效章节'
    ws.add_data_validation(dv_chapter)
    # Apply to column A (from row 3 to 500)
    
    # Data validation for 篇目
    dv_pianmu = DataValidation(type='list', formula1='"基础篇,综合篇,扩展篇"', allow_blank=True)
    dv_pianmu.error = '请选择基础篇、综合篇或扩展篇'
    ws.add_data_validation(dv_pianmu)
    
    # Data validation for 题型
    dv_tixing = DataValidation(type='list', formula1='"选择题,填空题,解答题"', allow_blank=True)
    ws.add_data_validation(dv_tixing)
    
    # Sample data
    samples = [
        ('第一章 函数、极限、连续', '基础篇', '选择题', '2026-06-18', 15, 14),
        ('第一章 函数、极限、连续', '综合篇', '选择题', '2026-06-18', 22, 21),
        ('第一章 函数、极限、连续', '基础篇', '解答题', '2026-06-18', 5, 5),
        ('第一章 函数、极限、连续', '综合篇', '解答题', '2026-06-18', 12, 6),
        ('第二章 一元函数微分学及其应用', '基础篇', '选择题', '2026-06-25', 17, 14),
        ('第二章 一元函数微分学及其应用', '综合篇', '选择题', '2026-06-25', 17, 14),
        ('第二章 一元函数微分学及其应用', '基础篇', '解答题', '2026-06-25', 17, 11),
        ('第二章 一元函数微分学及其应用', '综合篇', '解答题', '2026-06-25', 17, 13),
    ]
    
    for i, (ch, pm, tx, date, total, correct) in enumerate(samples):
        r = i + 3
        ws.cell(row=r, column=1, value=ch).font = data_font
        ws.cell(row=r, column=2, value=pm).font = data_font
        ws.cell(row=r, column=3, value=tx).font = data_font
        ws.cell(row=r, column=4, value=date).font = data_font
        ws.cell(row=r, column=5, value=total).font = data_font
        ws.cell(row=r, column=6, value=correct).font = data_font
        ws.cell(row=r, column=7).value = f'=IF(AND(E{r}<>"",F{r}<>""),F{r}/E{r},"")'
        ws.cell(row=r, column=7).number_format = '0.0%'
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply data validations to data rows
    dv_chapter.add(f'A3:A500')
    dv_pianmu.add(f'B3:B500')
    dv_tixing.add(f'C3:C500')
    
    # Conditional formatting: green-yellow-red color scale on 正确率
    ws.conditional_formatting.add('G3:G500',
        ColorScaleRule(start_type='min', start_color='FFC7CE',
                       mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                       end_type='max', end_color='C6EFCE'))
    
    # Freeze top row
    ws.freeze_panes = 'A3'
    
    return ws

# ============================================================
# Sheet 2: 数据汇总
# ============================================================
def create_summary_sheet(wb, record_ws):
    ws = wb.create_sheet(SUMMARY_SHEET)
    
    # Read record data
    records = []
    for r in range(3, record_ws.max_row + 1):
        ch = record_ws.cell(row=r, column=1).value
        pm = record_ws.cell(row=r, column=2).value
        tx = record_ws.cell(row=r, column=3).value
        total = record_ws.cell(row=r, column=5).value
        correct = record_ws.cell(row=r, column=6).value
        if ch and pm and tx and isinstance(total, (int, float)) and total > 0:
            records.append((str(ch), str(pm), str(tx), float(total), float(correct or 0)))
    
    if not records:
        ws.cell(row=1, column=1, value='暂无数据，请在刷题记录表中填入数据')
        return ws
    
    # Define column groups for summary
    # Each group: (label, filter_fn, col_total, col_rate)
    # filter_fn: (pianmu, tixing) -> bool
    col_groups = [
        ('基础-选择', lambda pm,tx: pm=='基础篇' and tx=='选择题'),
        ('基础-填空', lambda pm,tx: pm=='基础篇' and tx=='填空题'),
        ('基础-解答', lambda pm,tx: pm=='基础篇' and tx=='解答题'),
        ('基础-小计', lambda pm,tx: pm=='基础篇'),
        ('综合-选择', lambda pm,tx: pm=='综合篇' and tx=='选择题'),
        ('综合-填空', lambda pm,tx: pm=='综合篇' and tx=='填空题'),
        ('综合-解答', lambda pm,tx: pm=='综合篇' and tx=='解答题'),
        ('综合-小计', lambda pm,tx: pm=='综合篇'),
        ('扩展-解答', lambda pm,tx: pm=='扩展篇' and tx=='解答题'),
        ('扩展-小计', lambda pm,tx: pm=='扩展篇'),
        ('全章-选择', lambda pm,tx: tx=='选择题'),
        ('全章-填空', lambda pm,tx: tx=='填空题'),
        ('全章-解答', lambda pm,tx: tx=='解答题'),
        ('全章-选填合计', lambda pm,tx: tx in ('选择题','填空题')),
        ('全章-总计', lambda pm,tx: True),
    ]
    
    # Aggregate data by (chapter, group)
    agg = defaultdict(lambda: defaultdict(lambda: [0, 0]))  # ch -> group -> [total, correct]
    for ch, pm, tx, total, correct in records:
        for glabel, gfn in col_groups:
            if gfn(pm, tx):
                agg[ch][glabel][0] += total
                agg[ch][glabel][1] += correct
    
    # Get ordered chapter list (only chapters that have data)
    data_chapters = [ch for ch in CHAPTERS if ch in agg]
    if not data_chapters:
        data_chapters = sorted(agg.keys())
    
    # ===== Build Excel =====
    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_groups)*2+1)
    c = ws.cell(row=1, column=1, value='李林880 数据汇总')
    c.font = Font(bold=True, size=14)
    c.alignment = Alignment(horizontal='center')
    
    # Row 2: Column group headers (merged)
    n_cols = len(col_groups)
    col = 2  # start after chapter name column
    for glabel, _ in col_groups:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        c = ws.cell(row=2, column=col, value=glabel)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        col += 2
    
    ws.cell(row=2, column=1, value='章节').font = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # Row 3: Sub-headers (总题数 / 正确率)
    col = 2
    for _ in col_groups:
        c1 = ws.cell(row=3, column=col, value='总题数')
        c2 = ws.cell(row=3, column=col+1, value='正确率')
        for c in (c1, c2):
            c.font = header_font; c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
        col += 2
    
    ws.cell(row=3, column=1, value='').font = header_font
    ws.cell(row=3, column=1).fill = header_fill
    
    # Data rows
    row = 4
    for ch in data_chapters:
        ws.cell(row=row, column=1, value=ch).font = bold_font
        ws.cell(row=row, column=1).fill = section_fill
        
        col = 2
        for glabel, _ in col_groups:
            total, correct = agg[ch][glabel]
            c_total = ws.cell(row=row, column=col, value=int(total) if total > 0 else '')
            c_total.font = data_font
            c_total.alignment = Alignment(horizontal='center')
            
            c_rate = ws.cell(row=row, column=col+1)
            if total > 0:
                c_rate.value = correct / total
                c_rate.number_format = '0.0%'
            c_rate.font = data_font
            c_rate.alignment = Alignment(horizontal='center')
            
            col += 2
        
        # Borders
        for c in range(1, n_cols*2 + 2):
            ws.cell(row=row, column=c).border = border
        
        row += 1
    
    # Grand total row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    c = ws.cell(row=row, column=1, value='全部章节汇总')
    c.font = Font(bold=True, size=10, color='FFFFFFFF')
    c.fill = header_fill
    
    col = 2
    for glabel, _ in col_groups:
        total = sum(agg[ch][glabel][0] for ch in data_chapters)
        correct = sum(agg[ch][glabel][1] for ch in data_chapters)
        
        c_total = ws.cell(row=row, column=col, value=int(total) if total > 0 else '')
        c_total.font = Font(bold=True, size=10, color='FFFFFFFF')
        c_total.fill = header_fill
        c_total.alignment = Alignment(horizontal='center')
        
        c_rate = ws.cell(row=row, column=col+1)
        if total > 0:
            c_rate.value = correct / total
            c_rate.number_format = '0.0%'
        c_rate.font = Font(bold=True, size=10, color='FFFFFFFF')
        c_rate.fill = header_fill
        c_rate.alignment = Alignment(horizontal='center')
        
        col += 2
    
    for c in range(1, n_cols*2 + 2):
        ws.cell(row=row, column=c).border = border
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    for c in range(2, n_cols*2 + 2):
        ws.column_dimensions[get_column_letter(c)].width = 9
    
    # Conditional formatting for rate columns (all odd columns starting from col 3)
    rate_cols = []
    col = 3
    for _ in col_groups:
        rate_cols.append(get_column_letter(col))
        col += 2
    
    for rc in rate_cols:
        ws.conditional_formatting.add(f'{rc}4:{rc}{row}',
            ColorScaleRule(start_type='min', start_color='FFC7CE',
                           mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                           end_type='max', end_color='C6EFCE'))
    
    ws.freeze_panes = 'B4'
    
    return ws

# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    print('Loading 习题册.xlsx...')
    wb = openpyxl.load_workbook(XLSX_PATH)
    
    # Remove old 880 sheets if they exist
    for name in [RECORD_SHEET, SUMMARY_SHEET, '李林880']:
        if name in wb.sheetnames:
            del wb[name]
            print(f'  Removed old sheet: {name}')
    
    # Create sheets
    print('Creating 刷题记录 sheet...')
    record_ws = create_record_sheet(wb)
    
    print('Creating 数据汇总 sheet...')
    summary_ws = create_summary_sheet(wb, record_ws)
    
    # Save
    wb.save(XLSX_PATH)
    print(f'\nDone! Sheets added to {os.path.basename(XLSX_PATH)}')
    print(f'  - {RECORD_SHEET}')
    print(f'  - {SUMMARY_SHEET}')
