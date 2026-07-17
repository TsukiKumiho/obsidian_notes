import openpyxl, re
wb = openpyxl.load_workbook(r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx')

# Weight: typical exam point allocation per chapter
weights = {
    'DS': {1:3,2:9,3:5,4:2,5:11,6:7,7:7,8:5},  # 49 raw -> scaled to 45
    'CO': {1:3,2:7,3:11,4:5,5:11,6:4,7:6},        # 47 -> 45
    'OS': {1:4,2:13,3:9,4:6,5:5},                   # 37 -> 35
    'CN': {1:3,2:2,3:6,4:9,5:5,6:3},                # 28 -> 25
}
total_pts = {'DS':45, 'CO':45, 'OS':35, 'CN':25}

# Scale weights to actual point allocation
for subj in weights:
    w_sum = sum(weights[subj].values())
    for ch in weights[subj]:
        weights[subj][ch] = weights[subj][ch] / w_sum * total_pts[subj]

print('Weighted 408 Score (by exam frequency)')
print('='*55)

grand_w = 0
grand_u = 0

for idx, name in [(3,'DS'),(4,'CO'),(5,'OS'),(6,'CN')]:
    ws = wb.worksheets[idx]
    ch_data = {}
    ch_no = 0
    for r in range(5, ws.max_row+1):
        a = str(ws.cell(row=r, column=1).value or '')
        f = ws.cell(row=r, column=6).value or 0
        g = ws.cell(row=r, column=7).value or 0
        m = re.search(r'第(\d+)章', a)
        if m: ch_no = int(m.group(1)); continue
        if '合计' in a: break
        if isinstance(f,(int,float)) and float(f)>0:
            if ch_no not in ch_data: ch_data[ch_no] = [0,0]
            ch_data[ch_no][0] += float(f)
            ch_data[ch_no][1] += float(g or 0)
    
    # Compute weighted and unweighted
    subj_w = 0; subj_u = 0; n_ch = len(ch_data)
    print(f'\n{name} ({total_pts[name]}pts)')
    for ch_no in sorted(ch_data.keys()):
        t, c = ch_data[ch_no]
        rate = c/t*100 if t>0 else 0
        w = weights[name].get(ch_no, 0)
        subj_w += rate/100 * w
        subj_u += rate/100 * (total_pts[name]/n_ch)
        print(f'  Ch{ch_no}: {rate:5.1f}% x {w:4.1f}pts = {rate/100*w:5.1f}pts  ({int(c)}/{int(t)} zhenti)')
    print(f'  Weighted: {subj_w:.1f}  |  Unweighted(equal): {subj_u:.1f}')
    grand_w += subj_w
    grand_u += subj_u

print(f'\n{"="*55}')
print(f'Weighted total:   {grand_w:.1f} / 150')
print(f'Unweighted total: {grand_u:.1f} / 150')
print(f'Difference: {grand_w-grand_u:+.1f} pts')
