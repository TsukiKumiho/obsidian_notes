"""Generate 408 progress charts without pandas dependency."""
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']
plt.rcParams['axes.unicode_minus'] = False

SRC = r'C:\Users\34406\Documents\Obsidian Vault\习题册.xlsx'
OUT = r'C:\Users\34406\Documents\Obsidian Vault\.scripts\images'

wb = load_workbook(SRC)

def read_sheet(idx):
    """Read 408 sheet, return list of (section, total, correct, zhenti_total, zhenti_correct)"""
    ws = wb.worksheets[idx]
    rows = []
    in_data = False
    for r in range(1, ws.max_row+1):
        a = ws.cell(row=r, column=1).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value
        if a and '章' in str(a):
            in_data = True
            continue
        if in_data and a and isinstance(c, (int,float)) and c > 0:
            rows.append((str(a), float(c), float(d or 0), float(f or 0), float(g or 0)))
        if a and '合计' in str(a):
            break
    return rows

# Collect all data
all_data = {}
for idx, name in [(3,'DS'),(4,'CO'),(5,'OS'),(6,'CN')]:
    all_data[name] = read_sheet(idx)

# 1. 四科正确率总览 bar chart
fig, ax = plt.subplots(figsize=(8, 5))
names = []
rates = []
for name in ['DS','CO','OS','CN']:
    rows = all_data[name]
    total = sum(r[1] for r in rows)
    correct = sum(r[2] for r in rows)
    names.append(name)
    rates.append(correct/total*100 if total>0 else 0)

colors = ['#2F5496','#C55A11','#548235','#7030A0']
bars = ax.bar(names, rates, color=colors, edgecolor='white', linewidth=1.5)
for bar, rate in zip(bars, rates):
    ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.5, f'{rate:.1f}%',
            ha='center', va='bottom', fontsize=13, fontweight='bold')
ax.set_ylim(0, 100)
ax.set_ylabel('正确率 (%)', fontsize=12)
ax.set_title('408 四科一刷正确率总览', fontsize=15, fontweight='bold')
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f'{OUT}/408_overview.png', dpi=150, bbox_inches='tight')
plt.close()
print('1/5: 408 overview')

# 2. 计网章节明细 bar chart
fig, ax = plt.subplots(figsize=(10, 5))
cn_rows = all_data['CN']
labels = [r[0][:12] for r in cn_rows]
cn_rates = [r[2]/r[1]*100 if r[1]>0 else 0 for r in cn_rows]
bar_colors = ['#2F5496' if r>80 else '#C55A11' if r>70 else '#C00000' for r in cn_rates]
ax.barh(range(len(labels)), cn_rates, color=bar_colors, edgecolor='white')
ax.set_yticks(range(len(labels)))
ax.set_yticklabels(labels, fontsize=9)
ax.set_xlabel('正确率 (%)', fontsize=12)
ax.set_title('计网 各小节正确率', fontsize=14, fontweight='bold')
for i, (rate, row) in enumerate(zip(cn_rates, cn_rows)):
    ax.text(rate+0.5, i, f'{rate:.1f}% ({int(row[2])}/{int(row[1])})', va='center', fontsize=8)
ax.set_xlim(0, 110)
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
plt.tight_layout()
plt.savefig(f'{OUT}/cn_detail.png', dpi=150, bbox_inches='tight')
plt.close()
print('2/5: CN detail')

# 3. 计网真题vs自编
fig, ax = plt.subplots(figsize=(10, 5))
x = np.arange(len(labels))
w = 0.35
zt_rates = []
zb_rates = []
for r in cn_rows:
    zt_t, zt_c = r[3], r[4]
    zb_t, zb_c = r[1]-r[3], r[2]-r[4]
    zt_rates.append(zt_c/zt_t*100 if zt_t>0 else 0)
    zb_rates.append(zb_c/zb_t*100 if zb_t>0 else 0)

ax.bar(x-w/2, zt_rates, w, label='真题', color='#2F5496', edgecolor='white')
ax.bar(x+w/2, zb_rates, w, label='自编', color='#C55A11', edgecolor='white')
ax.set_xticks(x)
ax.set_xticklabels(labels, fontsize=8, rotation=30, ha='right')
ax.set_ylabel('正确率 (%)')
ax.set_title('计网 真题 vs 自编 对比', fontsize=14, fontweight='bold')
ax.legend(fontsize=11)
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f'{OUT}/cn_zt_vs_zb.png', dpi=150, bbox_inches='tight')
plt.close()
print('3/5: CN 真题vs自编')

# 4. 三科(DS/CO/OS) 真题vs自编 bar chart
fig, ax = plt.subplots(figsize=(8, 5))
x = np.arange(3)
w = 0.35
zt_all, zb_all = [], []
for name in ['DS','CO','OS']:
    rows = all_data[name]
    zt_t = sum(r[3] for r in rows)
    zt_c = sum(r[4] for r in rows)
    zb_t = sum(r[1]-r[3] for r in rows)
    zb_c = sum(r[2]-r[4] for r in rows)
    zt_all.append(zt_c/zt_t*100 if zt_t>0 else 0)
    zb_all.append(zb_c/zb_t*100 if zb_t>0 else 0)

ax.bar(x-w/2, zt_all, w, label='真题', color='#2F5496', edgecolor='white')
ax.bar(x+w/2, zb_all, w, label='自编', color='#C55A11', edgecolor='white')
for i in range(3):
    ax.text(i-w/2, zt_all[i]+0.5, f'{zt_all[i]:.1f}%', ha='center', fontsize=10, fontweight='bold')
    ax.text(i+w/2, zb_all[i]+0.5, f'{zb_all[i]:.1f}%', ha='center', fontsize=10, fontweight='bold')
ax.set_xticks(x)
ax.set_xticklabels(['DS','CO','OS'], fontsize=13)
ax.set_ylabel('正确率 (%)')
ax.set_title('三科 真题 vs 自编 对比', fontsize=14, fontweight='bold')
ax.legend(fontsize=11)
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.set_ylim(0, 100)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f'{OUT}/three_zt_zb.png', dpi=150, bbox_inches='tight')
plt.close()
print('4/5: 三科 真题vs自编')

# 5. 计网累计进度时间线
fig, ax = plt.subplots(figsize=(9, 4))
cn_dates = []
# Extract dates from sheet: B column
ws = wb.worksheets[6]
for r in range(5, 40):
    b = ws.cell(row=r, column=2).value
    c = ws.cell(row=r, column=3).value
    if b and isinstance(c, (int, float)) and c > 0:
        cn_dates.append((str(b), float(c)))

# Group by date
from collections import defaultdict
date_total = defaultdict(int)
date_label = {}
for d, t in cn_dates:
    date_total[d] += int(t)
    # Convert to readable date
    if len(d) == 3:  # e.g. "616" = June 16
        date_label[d] = f'6/{int(d[1:])}'

cumsum = 0
xs, ys, xlabels = [], [], []
for d in sorted(date_label.keys()):
    cumsum += date_total[d]
    xs.append(len(xs))
    ys.append(cumsum)
    xlabels.append(date_label[d])

ax.fill_between(xs, 0, ys, alpha=0.3, color='#2F5496')
ax.plot(xs, ys, 'o-', color='#2F5496', linewidth=2.5, markersize=8)
for i, y in enumerate(ys):
    ax.text(i, y+8, str(y), ha='center', fontsize=9, fontweight='bold')
ax.set_xticks(xs)
ax.set_xticklabels(xlabels, fontsize=9)
ax.set_ylabel('累计做题量', fontsize=12)
ax.set_title('计网 累计做题进度', fontsize=14, fontweight='bold')
ax.spines['top'].set_visible(False); ax.spines['right'].set_visible(False)
ax.yaxis.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig(f'{OUT}/cn_progress.png', dpi=150, bbox_inches='tight')
plt.close()
print('5/5: CN progress timeline')

print('All charts generated.')
